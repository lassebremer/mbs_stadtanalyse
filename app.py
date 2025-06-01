from flask import Flask, render_template, request, jsonify, send_file, Response, stream_with_context, url_for
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')  # Verwende nicht-interaktiven Backend
import matplotlib.pyplot as plt
import json
import io
import base64
import os
import re
import unicodedata # Für Umlaute etc. in simplified_name
from werkzeug.utils import secure_filename
import tempfile
from utils import (
    normalize, calculate_target_group, generate_charts,
    generate_scatter_plot, perform_clustering, encode_figure_to_base64,
    generate_city_pie_chart, generate_cities_chart,
    generate_interactive_scatter_plot, generate_interactive_clustering,
    generate_filtered_clustering, perform_clustering_population_target,
    generate_interactive_clustering_population_target
)
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans
from openpyxl import Workbook
import plotly.express as px
import sqlite3
from datetime import datetime
import time # Für eventuelle Pausen zwischen API-Aufrufen bei Paginierung
import requests # Für API-Aufrufe
import queue # Für SSE Statusmeldungen
from openpyxl.styles import Font, Alignment # Für Excel-Formatierung
from dotenv import load_dotenv # Für .env-Datei
import asyncio # Für asynchrone Verarbeitung
import aiohttp # Für asynchrone HTTP-Anfragen
import openpyxl # Stelle sicher, dass es importiert ist
from google.cloud import monitoring_v3
from google.oauth2 import service_account
from calendar import monthrange

# Lade Umgebungsvariablen aus der .env-Datei
load_dotenv()

# --- Bundesland Mapping ---
BUNDESLAND_MAP = {
    'BW': 'Baden-Württemberg',
    'BY': 'Bayern',
    'BE': 'Berlin',
    'BB': 'Brandenburg',
    'HB': 'Bremen',
    'HH': 'Hamburg',
    'HE': 'Hessen',
    'MV': 'Mecklenburg-Vorpommern',
    'NI': 'Niedersachsen',
    'NW': 'Nordrhein-Westfalen',
    'RP': 'Rheinland-Pfalz',
    'SL': 'Saarland',
    'SN': 'Sachsen',
    'ST': 'Sachsen-Anhalt',
    'SH': 'Schleswig-Holstein',
    'TH': 'Thüringen'
    # Füge ggf. weitere hinzu, falls andere Kürzel in der DB vorkommen
}
# --- Ende Bundesland Mapping ---

app = Flask(__name__)

# --- Datenbank-Hilfsfunktion: Spalte hinzufügen, falls nicht vorhanden ---
def add_column_if_not_exists(db_conn, table_name, column_name, column_type):
    cursor = db_conn.cursor()
    try:
        # PRAGMA table_info gibt Infos zu Spalten zurück
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = [column['name'] for column in cursor.fetchall()]
        if column_name not in columns:
            print(f"Füge Spalte '{column_name}' zur Tabelle '{table_name}' hinzu...")
            cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}")
            db_conn.commit()
            print(f"Spalte '{column_name}' erfolgreich hinzugefügt.")
        #     print(f"Spalte '{column_name}' existiert bereits in Tabelle '{table_name}'.")
    except sqlite3.Error as e:
        print(f"Fehler beim Überprüfen/Hinzufügen der Spalte '{column_name}' zu '{table_name}': {e}")
        # Es ist wichtig, hier keine Exception auszulösen, damit die App weiterläuft
    finally:
        # Cursor hier nicht schließen, da die Verbindung extern verwaltet wird
        pass

def get_db():
    db = sqlite3.connect(DATABASE)
    # Gib Zeilen zurück, die sich wie Dictionaries verhalten
    db.row_factory = sqlite3.Row
    return db

# --- Initialisierung beim App-Start (NACH get_db Definition) --- # NEUE POSITION
# Wird einmal ausgeführt, wenn die App startet (oder neu lädt im Debug-Modus)
with app.app_context():
    try:
        db = get_db()
        # Füge die postal_code Spalte zur place Tabelle hinzu, falls sie fehlt
        add_column_if_not_exists(db, 'place', 'postal_code', 'TEXT')
        # NEU: Spalten für search_cache hinzufügen
        add_column_if_not_exists(db, 'search_cache', 'min_rating', 'REAL')
        add_column_if_not_exists(db, 'search_cache', 'min_user_ratings', 'INTEGER')
        db.close() # Schließe die Verbindung nach der Prüfung/Änderung
    except Exception as e:
         print(f"Fehler während der Initialisierung der Datenbankstruktur: {e}")

# Konfiguration für Datei-Upload
UPLOAD_FOLDER = os.path.join(tempfile.gettempdir(), 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB Limit

# Lade Google Maps API-Schlüssel
DATABASE = 'data.db'
API_KEY = os.environ.get('GOOGLE_MAPS_API_KEY')

# Prüfe, ob der API-Schlüssel vorhanden ist
if not API_KEY:
    print("WARNUNG: Google Maps API-Schlüssel nicht gefunden. Bitte .env-Datei überprüfen.")
else:
    print(f"Google Maps API-Schlüssel geladen: {API_KEY[:4]}...{API_KEY[-4:]}")

# Globale Variable für SSE Nachrichten-Queue
search_status_queue = queue.Queue()


# --- Hilfsfunktion für Google Maps API ---
async def fetch_google_maps_data_async(session, API_KEY, query, city_id, city_display_name, fields_mask='*', latitude=None, longitude=None): 
    """Asynchrone Funktion für Google Maps Places searchText API Anfrage - mit optionalem locationBias für bessere lokale Ergebnisse."""
    url = 'https://places.googleapis.com/v1/places:searchText'
    logger = app.logger 

    headers = {
        'Content-Type': 'application/json',
        'X-Goog-Api-Key': API_KEY,
        'X-Goog-FieldMask': fields_mask 
    }
    
    # Erste Anfrage OHNE locationBias
    payload = {
        "textQuery": query,
        "languageCode": "de-DE",
        "maxResultCount": 20, 
        "regionCode": "de",
        "strictTypeFiltering": False,
        "rankPreference": "RELEVANCE"
    }

    logger.info(f"Google API: Rufe Ergebnisse für '{query}' ab (ohne Koordinaten)...")

    try:
        async with session.post(url, headers=headers, json=payload) as response:
            status_code = response.status
            if status_code == 200:
                result = await response.json()
                places = result.get("places", []) 
                
                # Wenn keine Ergebnisse gefunden wurden UND Koordinaten verfügbar sind, versuche es nochmal mit locationBias
                if len(places) == 0 and latitude is not None and longitude is not None:
                    logger.info(f"Keine Ergebnisse für '{query}' ohne Koordinaten. Versuche erneut mit locationBias...")
                    
                    # Zweite Anfrage MIT locationBias
                    payload_with_location = payload.copy()
                    payload_with_location["locationBias"] = {
                        "circle": {
                            "center": {
                                "latitude": latitude,
                                "longitude": longitude
                            },
                            "radius": 50000.0  # 50km Radius für Kreise/größere Gebiete
                        }
                    }
                    
                    async with session.post(url, headers=headers, json=payload_with_location) as response2:
                        status_code2 = response2.status
                        if status_code2 == 200:
                            result2 = await response2.json()
                            places = result2.get("places", [])
                            if len(places) > 0:
                                logger.info(f"Mit locationBias wurden {len(places)} Orte für '{query}' gefunden.")
                            else:
                                logger.info(f"Auch mit locationBias keine Ergebnisse für '{query}'.")
                        else:
                            logger.warning(f"Fehler beim Retry mit locationBias für '{query}': HTTP {status_code2}")
                else:
                    if len(places) > 0:
                        logger.info(f"Google API Erfolg für '{query}': {len(places)} Orte erhalten.")
                    else:
                        logger.info(f"Google API für '{query}': Keine Ergebnisse gefunden (0 Orte).")
                
                return {"places": places, "city_id": city_id, "city_display_name": city_display_name}
            else:
                error_text = await response.text()
                logger.error(f"Google Maps API Fehler für '{query}': HTTP {status_code}. Antwort: {error_text[:200]}...")
                return {"error": f"HTTP Error {status_code}", "status_code": status_code, "places": [], "city_id": city_id, "city_display_name": city_display_name}
    except Exception as e:
        logger.error(f"Google Maps API Exception für '{query}': {str(e)}", exc_info=True)
        return {"error": str(e), "places": [], "city_id": city_id, "city_display_name": city_display_name}

# Synchrone Version anpassen (analog)
def fetch_google_maps_data(API_KEY, query, fields_mask='*', latitude=None, longitude=None): 
    """Synchrone Funktion für Google Maps Places searchText API Anfrage - mit optionalem locationBias für bessere lokale Ergebnisse."""
    url = 'https://places.googleapis.com/v1/places:searchText'
    logger = app.logger
    headers = {
        'Content-Type': 'application/json',
        'X-Goog-Api-Key': API_KEY,
        'X-Goog-FieldMask': fields_mask 
    }
    
    # Erste Anfrage OHNE locationBias
    payload = {
        "textQuery": query,
        "languageCode": "de-DE",
        "maxResultCount": 20, 
        "regionCode": "de",
        "strictTypeFiltering": False,
        "rankPreference": "RELEVANCE"
    }

    logger.info(f"Google API (synchron): Rufe Ergebnisse für '{query}' ab (ohne Koordinaten)...")

    response = None
    try:
        response = requests.post(url, headers=headers, json=payload)
        response.raise_for_status()
        result = response.json()
        places = result.get("places", [])
        
        # Wenn keine Ergebnisse gefunden wurden UND Koordinaten verfügbar sind, versuche es nochmal mit locationBias
        if len(places) == 0 and latitude is not None and longitude is not None:
            logger.info(f"Keine Ergebnisse für '{query}' ohne Koordinaten. Versuche erneut mit locationBias...")
            
            # Zweite Anfrage MIT locationBias
            payload_with_location = payload.copy()
            payload_with_location["locationBias"] = {
                "circle": {
                    "center": {
                        "latitude": latitude,
                        "longitude": longitude
                    },
                    "radius": 50000.0  # 50km Radius für Kreise/größere Gebiete
                }
            }
            
            response2 = requests.post(url, headers=headers, json=payload_with_location)
            response2.raise_for_status()
            result2 = response2.json()
            places = result2.get("places", [])
            
            if len(places) > 0:
                logger.info(f"Mit locationBias wurden {len(places)} Orte für '{query}' gefunden.")
            else:
                logger.info(f"Auch mit locationBias keine Ergebnisse für '{query}'.")
        else:
            if len(places) > 0:
                logger.info(f"Google API Erfolg (synchron) für '{query}': {len(places)} Orte erhalten.")
            else:
                logger.info(f"Google API (synchron) für '{query}': Keine Ergebnisse gefunden (0 Orte).")
        
        return {"places": places}
    
    except requests.exceptions.RequestException as e:
        status_code = response.status_code if response else 500
        error_text = response.text if response else 'Keine Antwort'
        logger.error(f"Google Maps API Fehler (synchron) für '{query}': {e}. Status: {status_code}. Antwort: {error_text[:200]}...")
        return {"error": str(e), "status_code": status_code, "places": []} 
    except json.JSONDecodeError:
        logger.error(f"Fehler beim Parsen der JSON-Antwort (synchron) von Google Maps für '{query}'.")
        if response:
            logger.error(f"Antworttext (synchron): {response.text[:200]}...")
        return {"error": "Ungültige JSON-Antwort von der API (synchron)", "places": []}

async def save_places_to_db(db_path, term_id, results):
    """Asynchrone Funktion zum Speichern der Ergebnisse in der Datenbank."""
    # Datenbank-Verbindung aufbauen (in einem eigenen Thread)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    for result in results:
        if "error" in result:
            continue
            
        city_id = result["city_id"]
        city_display_name = result.get("city_display_name", f"Stadt ID {city_id}")
        found_places = result.get("places", [])
        
        # Explizit loggen, wenn keine Orte für eine Stadt gefunden wurden
        if len(found_places) == 0:
            print(f"Keine Orte zum Speichern für {city_display_name} (city_id: {city_id})")
            continue
        
        for place_data in found_places:
            place_id = place_data.get('id')
            if not place_id:
                continue

            # Aktuellen Zeitstempel generieren
            now_ts = datetime.now()
            
            # Extrahiere den displayName.text für den besseren Anzeigenamen
            display_name = None
            if place_data.get('displayName') and place_data['displayName'].get('text'):
                display_name = place_data['displayName']['text']

            # Extrahiere editorialSummary.text
            editorial_summary = None
            if place_data.get('editorialSummary') and place_data['editorialSummary'].get('text'):
                editorial_summary = place_data['editorialSummary']['text']

            # --- Place Tabelle --- # MODIFIZIERT: PLZ aus addressComponents holen
            postal_code_value = None
            address_components = place_data.get('addressComponents', [])
            if address_components:
                for component in address_components:
                    if 'postal_code' in component.get('types', []):
                        postal_code_value = component.get('longText') or component.get('shortText')
                        break # Nimm den ersten gefundenen

            place_values = {
                'place_id': place_id,
                'name': place_data.get('name'),
                'display_name': display_name,  # Speichere den displayName.text, falls vorhanden
                'formatted_address': place_data.get('formattedAddress'),
                'latitude': place_data.get('location', {}).get('latitude'),
                'longitude': place_data.get('location', {}).get('longitude'),
                'phone_number': place_data.get('internationalPhoneNumber'),
                'website_uri': place_data.get('websiteUri'),
                'google_maps_uri': place_data.get('googleMapsUri'),
                'price_level': place_data.get('priceLevel', '').split('$')[-1].count('€') if isinstance(place_data.get('priceLevel'), str) else None,
                'primary_type': place_data.get('types', [None])[0],
                'city_id': city_id,
                'last_updated': now_ts,
                'postal_code': postal_code_value, # NEUE SPALTE
                'supports_live_music': place_data.get('liveMusic'),
                'outdoor_seating': place_data.get('outdoorSeating'),
                'editorial_summary': editorial_summary
            }
            # Bereinige None-Werte
            place_values_clean = {k: v for k, v in place_values.items() if v is not None}

            # UPSERT Logik
            columns = ', '.join(place_values_clean.keys())
            placeholders = ', '.join(['?'] * len(place_values_clean))
            update_setters = ', '.join([f"{key} = excluded.{key}" for key in place_values_clean.keys() if key != 'place_id'])

            sql_place = f"""
                INSERT INTO place ({columns}) VALUES ({placeholders})
                ON CONFLICT(place_id) DO UPDATE SET {update_setters}
            """
            try:
                cursor.execute(sql_place, list(place_values_clean.values()))
            except sqlite3.Error as e:
                print(f"FEHLER beim Speichern von place {place_id}: {e}")
                continue

            # --- Place Type Tabelle ---
            types = place_data.get('types', [])
            for place_type in types:
                try:
                    cursor.execute(
                        "INSERT OR IGNORE INTO place_type (place_id, type) VALUES (?, ?)",
                        (place_id, place_type)
                    )
                except sqlite3.Error as e:
                    print(f"FEHLER beim Speichern von place_type für {place_id}: {e}")

            # --- Place Search Tabelle ---
            try:
                cursor.execute(
                    """
                    INSERT OR IGNORE INTO place_search (term_id, city_id, place_id, search_timestamp)
                    VALUES (?, ?, ?, ?)
                    """,
                    (term_id, city_id, place_id, now_ts)
                )
            except sqlite3.Error as e:
                print(f"FEHLER beim Speichern von place_search für {place_id} in {city_id}: {e}")

            # --- Rating History Tabelle ---
            rating = place_data.get('rating')
            user_rating_count = place_data.get('userRatingCount')
            if rating is not None and user_rating_count is not None:
                try:
                    cursor.execute(
                        """
                        INSERT INTO rating_history (place_id, rating, user_rating_count, timestamp)
                        VALUES (?, ?, ?, ?)
                        """,
                        (place_id, rating, user_rating_count, now_ts)
                    )
                except sqlite3.Error as e:
                    print(f"FEHLER beim Speichern der Bewertungshistorie für {place_id}: {e}")
    
            # --- Opening Hours Tabelle --- # NEUER ABSCHNITT
            opening_hours_data = place_data.get('regularOpeningHours')
            if opening_hours_data:
                weekday_text = "\n".join(opening_hours_data.get('weekdayDescriptions', []))
                periods_json = json.dumps(opening_hours_data.get('periods', []))
                try:
                    cursor.execute(
                        """
                        INSERT INTO opening_hours (place_id, weekday_text, periods_json)
                        VALUES (?, ?, ?)
                        ON CONFLICT(place_id) DO UPDATE SET
                            weekday_text = excluded.weekday_text,
                            periods_json = excluded.periods_json
                        """,
                        (place_id, weekday_text, periods_json)
                    )
                except sqlite3.Error as e:
                     print(f"FEHLER beim Speichern der Öffnungszeiten für {place_id}: {e}")

            # --- Reviews Tabelle --- # NEUER ABSCHNITT
            reviews_data = place_data.get('reviews', [])
            if reviews_data:
                for review in reviews_data:
                    publish_time_str = review.get('publishTime')
                    publish_time_dt = None
                    if publish_time_str:
                         try:
                            # Versuche, das ISO 8601 Format mit 'Z' zu parsen
                            publish_time_dt = datetime.fromisoformat(publish_time_str.replace('Z', '+00:00'))
                         except ValueError:
                              print(f"Warnung: Konnte publishTime '{publish_time_str}' für Review von {review.get('authorAttribution', {}).get('displayName')} nicht parsen.")

                    try:
                        cursor.execute(
                            """
                            INSERT OR IGNORE INTO review (
                                place_id, author_name, rating, 
                                relative_publish_time_description, text, language_code, publish_time
                            )
                            VALUES (?, ?, ?, ?, ?, ?, ?) 
                            """,
                            (
                                place_id,
                                review.get('authorAttribution', {}).get('displayName'),
                                review.get('rating'),
                                review.get('relativePublishTimeDescription'),
                                review.get('text', {}).get('text'),
                                review.get('text', {}).get('languageCode'),
                                publish_time_dt # Speichere als DATETIME
                            )
                        )
                    except sqlite3.Error as e:
                        print(f"FEHLER beim Speichern von Review für {place_id}: {e}")

    # Änderungen speichern und Verbindung schließen
    conn.commit()
    conn.close()

# --- Kernfunktion für die Massensuche ---
def run_place_search_for_all_cities(term_name, API_KEY, pause_between_cities=1):
    """Führt die Google Places Suche für einen Begriff über alle Städte durch und speichert Ergebnisse."""
    global search_status_queue
    
    # Hauptfunktion als asynchrone Funktion definieren
    async def run_search_async():
        
        logger = app.logger # Logger holen
        db = None 
        try:
            db = get_db()
            cursor = db.cursor()
            db_path = DATABASE  

            # Prüfen, ob der Suchbegriff bereits existiert
            cursor.execute("SELECT term_id FROM search_term WHERE name = ?", (term_name,))
            term_result = cursor.fetchone()
            
            # Wenn der Suchbegriff nicht existiert, füge ihn hinzu
            if not term_result:
                logger.info(f"Suchbegriff '{term_name}' noch nicht in der Datenbank, wird automatisch hinzugefügt.")
                search_status_queue.put(f"Neuer Suchbegriff '{term_name}' wird angelegt...")
                try:
                    cursor.execute("INSERT INTO search_term (name) VALUES (?)", (term_name,))
                    term_id = cursor.lastrowid
                    db.commit()
                    logger.info(f"Suchbegriff '{term_name}' (ID: {term_id}) zur Datenbank hinzugefügt.")
                    search_status_queue.put(f"Suchbegriff '{term_name}' erfolgreich angelegt.")
                except sqlite3.Error as e:
                    logger.error(f"Fehler beim Hinzufügen des Suchbegriffs '{term_name}': {e}")
                    search_status_queue.put(f"FEHLER: Suchbegriff '{term_name}' konnte nicht hinzugefügt werden: {e}")
                    search_status_queue.put("DONE")
                    db.close()
                    return
                
                # Suchbegriff nochmal aus DB holen, um ID zu bekommen
                cursor.execute("SELECT term_id FROM search_term WHERE name = ?", (term_name,))
                term_result = cursor.fetchone()
                
                if not term_result:
                    logger.error(f"Suchbegriff '{term_name}' konnte nicht in der Datenbank gefunden werden, obwohl er gerade hinzugefügt wurde.")
                    search_status_queue.put(f"FEHLER: Kritischer Datenbankfehler mit Suchbegriff '{term_name}'.")
                    search_status_queue.put("DONE")
                    db.close()
                    return
            
            term_id = term_result['term_id']
            logger.info(f"Starte Massensuche für Begriff: '{term_name}' (ID: {term_id})")
            search_status_queue.put(f"Starte Suche für '{term_name}'...")

            cursor.execute("SELECT city_id, name, simplified_name, latitude, longitude FROM city")
            cities = cursor.fetchall()
            total_cities = len(cities)
            logger.info(f"  -> {total_cities} Städte mit Koordinaten gefunden.")
            if total_cities == 0:
                logger.error("Keine Städte mit Koordinaten in der Datenbank für Massensuche gefunden.")
                search_status_queue.put("FEHLER: Keine Städte mit Koordinaten in der Datenbank gefunden.")
                return 
                
            search_status_queue.put(f"0/{total_cities} Städten verarbeitet.")

            rate_limit = 500  
            sem = asyncio.Semaphore(10)  # REDUZIERT VON 50
            
            processed_cities_count = 0 # Umbenannt für Klarheit
            total_places_found = 0
            start_time = time.time()
            batch_size = 50  
            
            async def process_city_with_rate_limit(city_data, session_context): # Parameter umbenannt
                # Verwende nonlocal um auf die Zählvariable im äußeren Scope zuzugreifen
                nonlocal processed_cities_count, total_places_found

                async with sem:
                    city_id_local = city_data['city_id']
                    city_query_name_local = city_data['simplified_name'] if city_data['simplified_name'] else city_data['name']
                    city_display_name_local = city_data['name']
                    # sqlite3.Row unterstützt kein .get(), deshalb try/except verwenden
                    try:
                        city_latitude = city_data['latitude']
                    except (KeyError, IndexError):
                        city_latitude = None
                    try:
                        city_longitude = city_data['longitude']
                    except (KeyError, IndexError):
                        city_longitude = None
                    
                    current_processing_num = processed_cities_count + 1 # Für Logging, bevor Zähler erhöht wird
                    query_text = f"{term_name} in {city_query_name_local}"
                    logger.info(f"Verarbeite Stadt: {city_display_name_local} ({current_processing_num}/{total_cities}) für Begriff '{term_name}'")
                    
                    api_result = await fetch_google_maps_data_async(
                        session_context, 
                        API_KEY, 
                        query_text, 
                        city_id_local, 
                        city_display_name_local,
                        latitude=city_latitude,
                        longitude=city_longitude
                    )
                    
                    processed_cities_count += 1 # Zähler hier erhöhen, nachdem der API-Call (potenziell lang) fertig ist

                    if "error" in api_result:
                        logger.error(f"API Fehler für {city_display_name_local}: {api_result.get('error')} (Status: {api_result.get('status_code', 'N/A')})")
                        search_status_queue.put(f"WARNUNG: API Fehler für {city_display_name_local} - {api_result.get('error')}")
                        # Gib das Fehlerobjekt weiter, damit es ggf. von save_places_to_db ignoriert wird
                        # oder die Plätze, die bis zum Fehler gesammelt wurden
                        return api_result 
                    
                    places_in_city_count = len(api_result.get("places", []))
                    total_places_found += places_in_city_count

                    elapsed_time = time.time() - start_time # Korrigierte Einrückung
                    estimated_remaining = (elapsed_time / processed_cities_count) * (total_cities - processed_cities_count) if processed_cities_count > 0 else 0
                    minutes, seconds = divmod(estimated_remaining, 60) # Korrigierte Einrückung
                        
                    status_message = f"{processed_cities_count}/{total_cities} Städte verarbeitet ({city_display_name_local}: {places_in_city_count} Orte). Geschätzte Restzeit: {int(minutes)} min {int(seconds)} sek."
                    search_status_queue.put(status_message)
                    logger.info(status_message)

                    delay_val = 60 / rate_limit  
                    await asyncio.sleep(delay_val)
                    
                    return api_result

            async with aiohttp.ClientSession() as session:
                for i in range(0, total_cities, batch_size):
                    batch_cities = cities[i:i+batch_size]
                    logger.info(f"Starte Verarbeitung für Batch {i//batch_size + 1} (Städte {i+1} bis {min(i+batch_size, total_cities)})")
                    
                    tasks_for_batch = [process_city_with_rate_limit(city_item, session) for city_item in batch_cities]
                    batch_api_results = []
                    try:
                        batch_api_results = await asyncio.gather(*tasks_for_batch, return_exceptions=True)
                    except Exception as gather_exc:
                        logger.error(f"Schwerwiegender Fehler während asyncio.gather für Batch {i//batch_size + 1}: {gather_exc}", exc_info=True)
                        search_status_queue.put(f"FEHLER: Kritischer Fehler bei der Verarbeitung eines Batches: {gather_exc}")
                        # Ggf. hier entscheiden, ob der gesamte Prozess abgebrochen werden soll

                    current_batch_places_to_save_db = []
                    for single_city_result in batch_api_results:
                        if isinstance(single_city_result, Exception):
                            logger.error(f"Exception in einem Task des Batches abgefangen: {single_city_result}", exc_info=True)
                            search_status_queue.put(f"WARNUNG: Fehler bei der Verarbeitung einer Stadt im Batch: {single_city_result}")
                            continue 
                        
                        # Auch wenn ein "error"-Key im Result ist (von fetch_google_maps_data_async so zurückgegeben),
                        # wollen wir ggf. die bis dahin gesammelten Orte speichern.
                        current_batch_places_to_save_db.append(single_city_result)
                    
                    if current_batch_places_to_save_db:
                        logger.info(f"Speichere {len(current_batch_places_to_save_db)} Ergebnisse (Städte) aus Batch {i//batch_size + 1} in der DB...")
                        await save_places_to_db(db_path, term_id, current_batch_places_to_save_db)
                        logger.info(f"Ergebnisse aus Batch {i//batch_size + 1} gespeichert.")
                    else:
                        logger.info(f"Keine Ergebnisse zum Speichern aus Batch {i//batch_size + 1}.")
                
                end_time = time.time()
                total_duration = end_time - start_time
                minutes, seconds = divmod(total_duration, 60)
                
                completion_message = f"Suche abgeschlossen. {total_places_found} Orte in {processed_cities_count} von {total_cities} Städten gefunden in {int(minutes)} min {int(seconds)} sek."
                logger.info(completion_message)
                search_status_queue.put(completion_message)
        
        except Exception as e:
            import traceback
            error_msg = f"Schwerwiegender Fehler im Hintergrund-Suchprozess: {e}"
            logger.error(f"{error_msg}\n{traceback.format_exc()}", exc_info=True)
            search_status_queue.put(f"FEHLER: {error_msg}")
            
        finally:
            search_status_queue.put("DONE")
            if db:
                db.close()
                logger.info("Datenbankverbindung im Hintergrundprozess geschlossen.")
    
    asyncio.run(run_search_async())
    return True

def process_data(min_age, max_age, w_pop, w_age, w_income):
    # --- Daten aus der Datenbank laden ---
    conn = None
    try:
        conn = get_db() # Verbindung zur data.db herstellen
        cursor = conn.cursor()

        # 1. Finde das aktuellste Jahr in den Demografie-Daten
        cursor.execute("SELECT MAX(year) FROM demographics")
        latest_year = cursor.fetchone()[0]
        if not latest_year:
            raise ValueError("Keine Demografie-Daten in der Datenbank gefunden.")

        # 2. Lade Stadt-, Demografie- und Altersverteilungsdaten für das aktuellste Jahr
        # Hauptabfrage für Stadt- und Demografiedaten, jetzt mit Modus (häufigster Wert) für Event/Gastro
        query_main = f'''
            WITH CityEventCounts AS (
                -- Zähle Vorkommen jedes Event-Textes pro Stadt
                SELECT
                    pc.city_id,
                    peg.event_gastro_text,
                    COUNT(peg.event_gastro_text) AS text_count
                FROM postal_code pc
                JOIN plz_event_gastro peg ON pc.postal_code_id = peg.postal_code_id
                -- Nur Texte berücksichtigen, die nicht leer oder NULL sind
                WHERE peg.event_gastro_text IS NOT NULL AND peg.event_gastro_text != \'\'
                GROUP BY pc.city_id, peg.event_gastro_text
            ),
            RankedCityEvents AS (
                -- Weise jedem Text pro Stadt einen Rang basierend auf der Häufigkeit zu
                SELECT
                    city_id,
                    event_gastro_text,
                    -- Rang 1 für den häufigsten. Bei Gleichstand wird der (alphabetisch) erste genommen.
                    ROW_NUMBER() OVER (PARTITION BY city_id ORDER BY text_count DESC, event_gastro_text ASC) AS rn
                FROM CityEventCounts
            )
            -- Hauptabfrage: Verbinde Stadt-/Demografie-Daten mit dem häufigsten Event-Text
            SELECT
                c.city_id,
                c.name AS location_name,
                c.simplified_name,
                c.bundesland AS Land,
                d.total_population AS total,
                d.income AS Einkommen_{latest_year},
                -- Hole den Text mit Rang 1 (den häufigsten) oder \'N/A\'
                COALESCE(rce.event_gastro_text, \'N/A\') AS event_gastro_text
            FROM city c
            JOIN demographics d ON c.city_id = d.city_id
            LEFT JOIN RankedCityEvents rce ON c.city_id = rce.city_id AND rce.rn = 1
            WHERE d.year = ?
            -- Kein GROUP BY c.city_id mehr nötig, da der Join nur den häufigsten Wert holt
        '''
        df_main = pd.read_sql_query(query_main, conn, params=(latest_year,))

        # Abfrage für Altersverteilungsdaten
        query_age = f"""
            SELECT
                d.city_id,
                ag.label AS age_group_label, -- Verwende Label für Pivot
                dad.count
            FROM demo_age_dist dad
            JOIN demographics d ON dad.demography_id = d.demography_id
            JOIN age_group ag ON dad.age_group_id = ag.age_group_id
            WHERE d.year = ?
        """
        df_age_dist = pd.read_sql_query(query_age, conn, params=(latest_year,))
        df_age_pivot = df_age_dist.pivot(index='city_id', columns='age_group_label', values='count').reset_index()
        df_age_pivot = df_age_pivot.fillna(0)

        # Zusammenführen der Hauptdaten mit den Altersdaten
        df = pd.merge(df_main, df_age_pivot, on='city_id', how='left')
        if not df.empty:
            if 'event_gastro_text' in df.columns:
                pass  # event_gastro_text ist vorhanden

        # Einkommensspalte umbenennen, falls nötig, damit sie generisch 'Einkommen' heißt
        # Der Rest des Codes erwartet 'Einkommen_2022'
        # Wir benennen die dynamische Spalte in die erwartete um
        # Vorsicht: Wenn latest_year nicht 2022 ist, könnte dies irreführend sein.
        # Besser wäre es, den restlichen Code anzupassen, aber für minimale Änderungen machen wir es so:
        if f'Einkommen_{latest_year}' in df.columns and 'Einkommen_2022' not in df.columns:
            df = df.rename(columns={f'Einkommen_{latest_year}': 'Einkommen_2022'})
        elif 'Einkommen_2022' not in df.columns:
            # Fallback, falls die Spalte fehlt (sollte nicht passieren)
            print(f"Warnung: Erwartete Einkommensspalte 'Einkommen_{latest_year}' oder 'Einkommen_2022' nicht im DataFrame.")
            df['Einkommen_2022'] = 0 # Standardwert setzen

        # --- Bundesland-Abkürzungen durch volle Namen ersetzen ---
        if 'Land' in df.columns:
             # .map() ersetzt Werte basierend auf dem Dictionary
             # .fillna(df['Land']) behält den Originalwert (Abkürzung), falls keine Übereinstimmung gefunden wird
             df['Land'] = df['Land'].str.strip() # Entferne führende/nachgestellte Leerzeichen
             df['Land'] = df['Land'].map(BUNDESLAND_MAP).fillna(df['Land'])
             # Optional: Falls Nicht-Gemappte zu 'N/A' werden sollen:
             # df['Land'] = df['Land'].map(BUNDESLAND_MAP).fillna('N/A')
        else:
             print("Warnung: Spalte 'Land' (Bundesland) nicht im DataFrame nach SQL-Abfrage gefunden.")
        # --- Ende Bundesland Mapping ---

    except sqlite3.Error as e:
        print(f"Datenbankfehler beim Laden der Daten: {e}")
        # Hier könnte man eine leere Tabelle zurückgeben oder einen Fehler werfen
        return {'error': f"Datenbankfehler: {e}"}
    finally:
        if conn:
            conn.close()

    # --- Datenverarbeitung (ab hier weitgehend wie vorher) ---

    # Stelle sicher, dass simplified_name existiert (sollte aus DB kommen)
    if 'simplified_name' not in df.columns:
        print("Warnung: Spalte 'simplified_name' nicht in Datenbank gefunden. Verwende 'location_name'.")
        df['simplified_name'] = df['location_name']
    # Fülle fehlende simplified_name mit location_name als Fallback
    df['simplified_name'] = df['simplified_name'].fillna(df['location_name'])

    # Deutschland-Eintrag entfernen (sollte in DB nicht vorkommen)

    # Bundesland-Mapping entfernen (kommt direkt aus DB)

    # Zielgruppe berechnen
    # WICHTIG: calculate_target_group benötigt die Altersspalten aus df_age_pivot.
    # Stelle sicher, dass die Spaltennamen in df_age_pivot mit denen übereinstimmen,
    # die calculate_target_group erwartet (z.B. '0-17', '18-24', ...)
    # Falls calculate_target_group in utils.py angepasst werden muss, ist das ein separater Schritt.
    try:
        df = calculate_target_group(df, min_age, max_age)
    except KeyError as e:
        print(f"Fehler in calculate_target_group: Fehlende Spalte {e}. Überprüfe die Altersspaltennamen im DataFrame und in utils.py.")
        return {'error': f"Fehler bei der Zielgruppenberechnung: Spalte {e} fehlt."}
    except Exception as e:
        print(f"Unbekannter Fehler in calculate_target_group: {e}")
        return {'error': f"Fehler bei der Zielgruppenberechnung: {e}"}

    # Normalisierung (verwende 'Einkommen_2022')
    # Sicherstellen, dass die Spalten existieren und numerisch sind
    for col in ['total', 'Einkommen_2022', 'target_group_percent']:
        if col not in df.columns:
             print(f"Warnung: Erwartete Spalte '{col}' für Normalisierung nicht gefunden.")
             # Ggf. Fehler werfen oder Standardwert setzen
             df[col] = 0 # Beispiel: Standardwert setzen
        # Konvertiere zu numerischem Typ, falls möglich, behandle Fehler
        df[col] = pd.to_numeric(df[col], errors='coerce')
        # Fülle NaN-Werte, die durch Konvertierungsfehler entstanden sein könnten
        if df[col].isnull().any():
            print(f"Warnung: NaN-Werte in Spalte '{col}' nach Konvertierung gefunden. Werden mit 0 gefüllt.")
            df[col] = df[col].fillna(0)

    # Min/Max Werte für Normalisierung berechnen (nachdem NaN entfernt wurde)
    total_min, total_max = df["total"].min(), df["total"].max()
    income_min, income_max = df["Einkommen_2022"].min(), df["Einkommen_2022"].max()
    target_min, target_max = df["target_group_percent"].min(), df["target_group_percent"].max()

    # Normalisierungsfunktion anwenden
    df["norm_pop"] = df["total"].apply(lambda x: normalize(x, total_min, total_max) if total_max > total_min else 0)
    df["norm_income"] = df["Einkommen_2022"].apply(lambda x: normalize(x, income_min, income_max) if income_max > income_min else 0)
    df["norm_target"] = df["target_group_percent"].apply(lambda x: normalize(x, target_min, target_max) if target_max > target_min else 0)

    # Gewichte normalisieren
    total_weight = w_pop + w_age + w_income
    w_pop /= total_weight
    w_age /= total_weight
    w_income /= total_weight

    # Score berechnen
    df["score"] = df["norm_pop"] * w_pop + df["norm_income"] * w_income + df["norm_target"] * w_age

    # Clustering durchführen, um Cluster-Namen zu erhalten
    # Feature-Auswahl für Clustering
    features_for_clustering = df[['total', 'Einkommen_2022', 'target_group_percent']].copy()
    # Sicherstellen, dass keine NaN-Werte im Clustering sind
    features_for_clustering = features_for_clustering.fillna(0)
    
    # Skalierung der Daten
    scaler = StandardScaler()
    scaled_features = scaler.fit_transform(features_for_clustering)
    
    # K-Means Clustering
    kmeans = KMeans(n_clusters=5, random_state=42)
    raw_clusters = kmeans.fit_predict(scaled_features)
    
    # Zentren berechnen
    centers = scaler.inverse_transform(kmeans.cluster_centers_)
    
    # Zuordnung der Cluster zu den Interpretationen
    cluster_mapping = {}
    
    # Finde den Cluster mit den höchsten Bevölkerungszahlen -> Großstädte (blau, Cluster 2)
    cluster_mapping[np.argmax([center[0] for center in centers])] = 1  # Blau
    
    # Finde den Cluster mit dem höchsten Zielgruppenanteil -> Universitätsstädte (lila, Cluster 4)
    cluster_mapping[np.argmax([center[2] for center in centers])] = 3  # Lila
    
    # Finde den Cluster mit dem höchsten Einkommen -> Wohlhabende Mittelstädte (orange, Cluster 5)
    cluster_mapping[np.argmax([center[1] for center in centers])] = 4  # Orange
    
    # Finde den Cluster mit dem niedrigsten Einkommen -> Ländliche Regionen (grün, Cluster 3)
    cluster_mapping[np.argmin([center[1] for center in centers])] = 2  # Grün
    
    # Der übrig gebliebene Cluster ist Mittelstädte (rot, Cluster 1)
    for i in range(5):
        if i not in cluster_mapping:
            cluster_mapping[i] = 0  # Rot
    
    # Cluster-Namen definieren
    cluster_names = {
        0: "Mittelstädte",
        1: "Großstädte",
        2: "Ländliche Regionen",
        3: "Universitätsstädte",
        4: "Wohlhabende Mittelstädte"
    }
    
    # Cluster-Zuweisungen und Namen zum Dataframe hinzufügen
    df['cluster_id'] = [cluster_mapping[c] for c in raw_clusters]
    df['cluster_name'] = df['cluster_id'].map(cluster_names)

    # Stelle sicher, dass die relevanten Namen im result DataFrame sind
    # Verwende city_id statt dem alten Hash für die ID
    result = df[["city_id", "location_name", "simplified_name", "Land", "total", "Einkommen_2022", "target_group_percent", "cluster_id", "cluster_name", "score", "event_gastro_text"]].copy() # Neue Spalte hinzugefügt
    
    # Ausgabe vorbereiten - die Spalten under_18_percent und others_percent ausblenden
    result["target_group_percent"] = (result["target_group_percent"] * 100).round(1).astype(str) + "%"
    
    # Formatiere Zahlen mit Tausendertrennzeichen (Punkt)
    result["total"] = result["total"].apply(lambda x: f"{int(x):,}".replace(",", "."))
    result["Einkommen_2022"] = result["Einkommen_2022"].apply(lambda x: f"{int(x):,}".replace(",", "."))
    
    # Score numerisch halten für korrekte Sortierung
    result["score"] = result["score"].round(4)

    # Balken + Kreisdiagramm
    # generate_charts erwartet Spaltennamen wie im Original-df
    # df sollte jetzt kompatibel sein.
    bar_chart_top10, sorted_cities, pie_chart = generate_charts(df)
    
    # Diagramme als base64 kodieren
    bar_img_top10 = io.BytesIO()
    bar_chart_top10.savefig(bar_img_top10, format='png', bbox_inches='tight')
    bar_img_top10.seek(0)
    bar_encoded_top10 = base64.b64encode(bar_img_top10.getvalue()).decode('utf-8')
    
    # Figur schließen, um Speicher freizugeben
    plt.close(bar_chart_top10)
    
    # Finde den minimalen und maximalen Score-Wert für einheitliche Skalierung
    min_score = sorted_cities['score'].min()
    max_score = sorted_cities['score'].max()
    x_min = max(0, min_score - (max_score - min_score) * 0.1)
    
    # Generiere weitere Plots für Städtegruppen (bis zu 5 Seiten)
    more_charts = []
    total_cities = len(sorted_cities)
    max_pages = min(5, (total_cities + 9) // 10)  # Bis zu 5 Seiten oder alle verfügbaren Städte
    
    for page in range(1, max_pages):
        start_idx = page * 10
        end_idx = min(start_idx + 10, total_cities)
        if start_idx < total_cities:
            # Erstelle Diagramm für diese Gruppe - mit einheitlicher Skalierung
            page_chart = generate_cities_chart(
                sorted_cities.iloc[start_idx:end_idx], 
                start_idx, 
                x_min=x_min,
                global_max=max_score
            )
            
            # Als base64 kodieren
            chart_img = io.BytesIO()
            page_chart.savefig(chart_img, format='png', bbox_inches='tight')
            chart_img.seek(0)
            chart_encoded = base64.b64encode(chart_img.getvalue()).decode('utf-8')
            
            # Figur schließen, um Speicher freizugeben
            plt.close(page_chart)
            
            # Zum Ergebnis hinzufügen
            more_charts.append({
                'start': start_idx + 1,
                'end': end_idx,
                'chart': chart_encoded
            })
    
    pie_img = io.BytesIO()
    pie_chart.savefig(pie_img, format='png', bbox_inches='tight')
    pie_img.seek(0)
    pie_encoded = base64.b64encode(pie_img.getvalue()).decode('utf-8')
    
    # Figur schließen, um Speicher freizugeben
    plt.close(pie_chart)

    # Extrahiere die Daten für die Frontend-Verarbeitung
    cities_data_for_charts = sorted_cities[['location_name', 'score']].copy()
    cities_data_for_charts['rank'] = range(1, len(cities_data_for_charts) + 1)
    cities_data_for_charts = cities_data_for_charts.to_dict(orient='records')
    
    # Tabellendaten als HTML mit angepasstem Styling
    result_sorted = result.sort_values("score", ascending=False).reset_index(drop=True)
    
    # Platzierungsspalte hinzufügen (von 1 beginnend)
    result_sorted.index = result_sorted.index + 1
    result_sorted = result_sorted.reset_index().rename(columns={'index': 'Platz'})
    
    # ID-Spalte hinzufügen (jetzt direkt city_id verwenden)
    result_sorted['id'] = result_sorted['city_id'] # Verwende die echte city_id
    
    # Spaltennamen für die Anzeige anpassen
    column_labels = {
        'city_id': 'id',
        'location_name': 'Stadt',
        'Land': 'Bundesland',
        'total': 'Einwohner',
        'Einkommen_2022': 'Einkommen',
        'target_group_percent': 'Zielgruppe',
        'cluster_name': 'Cluster',
        'score': 'Score',
        'event_gastro_text': 'Vertriebsnummer' # Neuer Name hier
    }
    
    # Nur die Anzeigespalten umbenennen und city_id behalten für den formatter
    result_sorted_display = result_sorted.rename(columns=column_labels)

    # Spalten bestimmen, die angezeigt werden sollen (id, simplified_name, cluster_id ausschließen)
    # Füge die neue Spalte hinzu
    display_columns = [col for col in result_sorted_display.columns if col not in ['simplified_name', 'cluster_id'] and col != 'id']

    # Stadtstaaten erkennen
    stadtstaaten = ['Berlin', 'Bremen', 'Hamburg']
    
    # Benutzerdefinierte Konvertierung für HTML-Tabelle mit CSS-Klassen
    def html_formatter(df_display, df_original): # Übergebe beide DataFrames
        # HTML-Tabelle erstellen
        html = '<table id="data-table" class="table table-striped table-hover">\n'
        
        # Tabellenkopf
        html += '<thead>\n<tr>'
        html += '<th class="select-all-checkbox-header"><input type="checkbox" class="select-all-checkbox" title="Alle auswählen"></th>'
        
        # Bestehende Überschriften mit CSS-Klassen
        # Passe die Spaltenliste an, da 'id' nicht angezeigt werden soll
        display_columns_local = [col for col in df_display.columns if col not in ['id', 'simplified_name', 'cluster_id']]
        # Reihenfolge festlegen (optional, aber schöner)
        column_order = ['Platz', 'Stadt', 'Bundesland', 'Einwohner', 'Einkommen', 'Zielgruppe', 'Score', 'Cluster', 'Vertriebsnummer'] # Neuer Name hier
        display_columns_local = [col for col in column_order if col in df_display.columns]

        for col in display_columns_local:
            css_class = ""
            if col == "Platz":
                css_class = "text-center"
            elif col == "Stadt":
                css_class = "text-left"
            elif col == "Bundesland":
                css_class = "text-center"
            elif col in ["Einwohner", "Einkommen"]: # Zielgruppe und Score getrennt behandeln
                css_class = "text-right"
            elif col == "Zielgruppe":
                 css_class = "text-right"
            elif col == "Score":
                 css_class = "text-right"
            elif col == "Cluster":
                css_class = "text-center"
            elif col == "Vertriebsnummer": # Neuer Name hier und CSS-Klasse
                 css_class = "text-left vertriebsnummer-cell"

            html += f'<th class="{css_class}">{col}</th>'
        
        # html += '<th class="text-center">Lokale Suche</th>'
        
        html += '</tr>\n</thead>\n'
        
        # Tabelleninhalt
        html += '<tbody>\n'
        
        # Füge Zeilen hinzu - iteriere über den Original-DataFrame für einfachen Zugriff auf alle Spalten
        for idx, row_original in df_original.iterrows(): # Verwende df_original (enthält city_id etc.)
            city_id = row_original['city_id'] # Hole die echte city_id
            city_display_name = row_original['location_name']
            city_search_name = row_original['simplified_name']
            # Finde die entsprechende Zeile im Display-DataFrame über den Index
            row_display = df_display.loc[idx]
            
            # *** Daten für den Button holen ***
            html += f'<tr class="city-row" data-id="{city_id}" data-city-search="{city_search_name}" data-city-display="{city_display_name}">'
            
            # Checkbox-Zelle
            html += '<td class="text-center"><input type="checkbox" class="select-checkbox" title="Auswählen"></td>'
            
            # Restliche Spalten mit CSS-Klassen - verwende row_display für formatierte Werte
            for col in display_columns_local:
                css_class = ""
                cell_content = row_display[col] # Hole den (potenziell formatierten) Wert aus df_display
                
                # Bestimme CSS-Klasse basierend auf Spaltenname
                if col == "Platz":
                    css_class = "text-center font-weight-bold"
                elif col == "Stadt":
                    css_class = "text-left"
                    cell_content = str(cell_content)
                elif col == "Bundesland":
                    css_class = "text-center bundesland-cell"
                    if cell_content in stadtstaaten:
                        css_class += " stadtstaaten"
                elif col in ["Einwohner", "Einkommen"]:
                    css_class = "text-right"
                    # Konvertierung schon vorher passiert, nur formatieren
                    if isinstance(cell_content, (str)) and '.' in cell_content: # Bereits formatiert?
                       pass
                    elif isinstance(cell_content, (int, float, np.integer, np.floating)):
                       cell_content = f'{int(cell_content):,}'.replace(',', '.')
                    else: # Fallback, falls Konvertierung fehlschlug
                       cell_content = str(cell_content)
                elif col == "Zielgruppe":
                    css_class = "text-right"
                    # Sollte schon formatiert sein
                    if not isinstance(cell_content, str) or '%' not in cell_content:
                        try:
                            cell_content = f'{float(cell_content):.1f}%'
                        except (ValueError, TypeError):
                            cell_content = 'N/A'
                elif col == "Score":
                    css_class = "text-right"
                    # Score ist numerisch für Sortierung, jetzt formatieren
                    if isinstance(cell_content, (float, np.floating)):
                        try:
                            cell_content = f'{(cell_content * 100):.2f}%' # Zeige als Prozent
                        except (ValueError, TypeError):
                             cell_content = 'N/A'
                    else: # Fallback
                         cell_content = str(cell_content)
                elif col == "Cluster":
                    css_class = "text-center"
                    # Nutze die cluster_id aus dem Original-DataFrame für die Klasse
                    cluster_id_from_db = row_original['cluster_id'] # Ist 0-basiert
                    cluster_name_text = str(cell_content) # Der Name aus dem Display-df
                    # Konvertiere 0-basierte ID in 1-basierte Klasse für CSS
                    css_cluster_num = cluster_id_from_db + 1 
                    cell_content = f'<span class="cluster-badge cluster-{css_cluster_num}">{cluster_name_text}</span>'
                elif col == "Vertriebsnummer": # Neuer Name hier
                     css_class = "text-left vertriebsnummer-cell"
                     cell_content = str(cell_content) if pd.notna(cell_content) and str(cell_content).strip() != 'N/A' else '' # Zeige 'N/A' nicht an, sondern leer

                html += f'<td class="{css_class}">{cell_content}</td>'
            
            # html += f'<td class="text-center"><button class="btn btn-sm btn-outline-primary open-search-popup" data-city="{city_search_name}" data-city-display="{city_display_name}" title="Lokale Suche für {city_display_name} öffnen"><i class="fas fa-search-location"></i></button></td>'
            
            html += '</tr>\n'
        
        html += '</tbody>\n</table>'
        
        return html
    
    # Stelle sicher, dass die benötigten Spalten für den Formatter existieren
    # df_original ist result_sorted, df_display ist result_sorted_display
    table_html = html_formatter(result_sorted_display, result_sorted)
    
    # Speichere das vollständige DataFrame für die Detailansicht
    # df enthält jetzt die kombinierten Daten aus der DB
    cities_data = df.copy()
    
    # Erweiterte Analysen
    # Stellen sicher, dass die benötigten Spalten numerisch sind
    df_analysis = df.copy()
    for col in ['total', 'Einkommen_2022', 'target_group_percent', 'score', 'norm_pop', 'norm_income', 'norm_target']:
        if col in df_analysis.columns:
            df_analysis[col] = pd.to_numeric(df_analysis[col], errors='coerce').fillna(0)
        else:
             print(f"Warnung: Spalte '{col}' für erweiterte Analyse nicht gefunden.")
             df_analysis[col] = 0 # Fallback

    # 1. Scatter Plot
    scatter_fig = generate_scatter_plot(df_analysis)
    scatter_encoded = encode_figure_to_base64(scatter_fig)
    plt.close(scatter_fig)
    
    # 1b. Interaktiver Scatter Plot mit Plotly
    interactive_scatter = generate_interactive_scatter_plot(df_analysis)
    
    # 2. Clustering-Analyse (Einwohner vs. Einkommen)
    cluster_fig, cluster_summary = perform_clustering(df_analysis, n_clusters=5)
    cluster_encoded = encode_figure_to_base64(cluster_fig)
    plt.close(cluster_fig)
    
    cluster_table = cluster_summary.to_html(
        classes='table table-striped table-hover',
        index=False,
        justify='left',
        table_id='cluster-table'
    )
    
    # 2b. Interaktives Clustering mit Plotly (mit 5 Clustern)
    interactive_clustering = generate_interactive_clustering(df_analysis, n_clusters=5)
    
    # 3. Clustering-Analyse (Einwohner vs. Zielgruppe)
    cluster2_fig, cluster2_summary = perform_clustering_population_target(df_analysis, n_clusters=5)
    cluster2_encoded = encode_figure_to_base64(cluster2_fig)
    plt.close(cluster2_fig)
    
    cluster2_table = cluster2_summary.to_html(
        classes='table table-striped table-hover',
        index=False,
        justify='left',
        table_id='cluster2-table'
    )
    
    # 3b. Interaktives Clustering 2 mit Plotly (Einwohner vs. Zielgruppe, mit 5 Clustern)
    interactive_clustering2 = generate_interactive_clustering_population_target(df_analysis, n_clusters=5)
    
    # NaN-Werte durch None ersetzen, um gültiges JSON zu gewährleisten
    # Verwende df_analysis oder df, je nachdem, was die vollständigen Daten enthält
    cities_data_records = df_analysis.replace({np.nan: None}).to_dict(orient='records')
    
    return {
        'table_html': table_html,
        'bar_chart_top10': bar_encoded_top10,
        'more_charts': more_charts,
        'cities_for_charts': cities_data_for_charts,
        'pie_chart': pie_encoded,
        'advanced_analysis': {
            'scatter_plot': scatter_encoded,
            'interactive_scatter': interactive_scatter,
            'clustering': {
                'plot': cluster_encoded,
                'table': cluster_table
            },
            'interactive_clustering': interactive_clustering,
            'clustering2': {
                'plot': cluster2_encoded,
                'table': cluster2_table
            },
            'interactive_clustering2': interactive_clustering2
        },
        'cities_data': cities_data_records
    }

@app.route('/')
def index():
    # Nach dem Starten der App: Aktualisiere display_name für bestehende Einträge falls nötig
    update_display_names()
    return render_template('index.html')
    
def update_display_names():
    """Aktualisiert display_name für bestehende Einträge, wenn sie leer sind."""
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Prüfe, ob die Spalte existiert
        cursor.execute("PRAGMA table_info(place)")
        columns = cursor.fetchall()
        column_names = [col['name'] for col in columns]
        
        if 'display_name' in column_names:
            # Hole alle Orte ohne display_name
            cursor.execute("SELECT place_id, name FROM place WHERE display_name IS NULL")
            places = cursor.fetchall()
            
            if places:
                print(f"Aktualisiere display_name für {len(places)} Orte...")
                for place in places:
                    try:
                        cursor.execute(
                            "UPDATE place SET display_name = ? WHERE place_id = ?",
                            (place['name'], place['place_id'])
                        )
                    except sqlite3.Error as e:
                        print(f"Fehler beim Aktualisieren von display_name für {place['place_id']}: {e}")
                
                db.commit()
                print("Display-Namen aktualisiert.")
        
        db.close()
    except Exception as e:
        print(f"Fehler beim Aktualisieren der display_name: {e}")
        if 'db' in locals():
            db.close()

@app.route('/city_chart/<int:city_id>', methods=['GET']) # Verwende city_id als Integer
def city_chart(city_id):
    """Generiert ein Kreisdiagramm für eine bestimmte Stadt aus der Datenbank"""
    conn = None
    try:
        conn = get_db() # Verbindung zur data.db
        cursor = conn.cursor()

        # Finde das aktuellste Jahr
        cursor.execute("SELECT MAX(year) FROM demographics")
        latest_year = cursor.fetchone()[0]
        if not latest_year:
            return jsonify({'error': 'Keine Demografie-Daten gefunden'}), 500

        # Lade die benötigten Daten für die spezifische Stadt
        query_main = f"""
            SELECT
                c.city_id,
                c.name AS location_name,
                d.total_population AS total
            FROM city c
            JOIN demographics d ON c.city_id = d.city_id
            WHERE c.city_id = ? AND d.year = ?
        """
        city_main_data = pd.read_sql_query(query_main, conn, params=(city_id, latest_year))

        if city_main_data.empty:
            return jsonify({'error': 'Stadt nicht gefunden oder keine Demografiedaten für das Jahr'}), 404

        # Lade Altersverteilungsdaten für die Stadt
        query_age = f"""
            SELECT
                ag.label AS age_group_label,
                dad.count
            FROM demo_age_dist dad
            JOIN demographics d ON dad.demography_id = d.demography_id
            JOIN age_group ag ON dad.age_group_id = ag.age_group_id
            WHERE d.city_id = ? AND d.year = ?
        """
        df_age_dist = pd.read_sql_query(query_age, conn, params=(city_id, latest_year))

        # Erstelle einen DataFrame für diese eine Stadt mit den Altersspalten
        city_df = city_main_data.copy()
        for _, row in df_age_dist.iterrows():
            city_df[row['age_group_label']] = row['count']
        city_df = city_df.fillna(0)

        # Aktuelle Parameter für Zielgruppe holen (aus Request oder Standard)
        min_age = int(request.args.get('min_age', 18))
        max_age = int(request.args.get('max_age', 35))

        # Zielgruppe für diese Stadt berechnen
        # calculate_target_group erwartet einen DataFrame, also übergeben wir city_df
        city_data_processed = calculate_target_group(city_df, min_age, max_age)

        if city_data_processed.empty:
             return jsonify({'error': 'Fehler bei der Zielgruppenberechnung für die Stadt'}), 500

        # Die erste (und einzige) Zeile enthält die benötigten Prozente
        city_row = city_data_processed.iloc[0]
        
        # Prüfe, ob die benötigten Prozent-Spalten existieren
        required_pie_cols = ['under_18_percent', 'target_group_percent', 'others_percent']
        if not all(col in city_row for col in required_pie_cols):
             print(f"WARNUNG: Fehlende Spalten für Kreisdiagramm: { {col for col in required_pie_cols if col not in city_row} }")
             # Fallback oder Fehlermeldung
             return jsonify({'error': 'Daten für Kreisdiagramm unvollständig'}), 500

        # Kreisdiagramm erstellen
        pie_chart = generate_city_pie_chart(city_row)
        
        # Als base64 kodieren
        pie_encoded = encode_figure_to_base64(pie_chart)
        
        # Figur schließen, um Speicher freizugeben
        plt.close(pie_chart)
        
        return jsonify({
            'pie_chart': pie_encoded,
            'city_name': city_row['location_name']
        })

    except sqlite3.Error as e:
        print(f"Datenbankfehler in /city_chart/{city_id}: {e}")
        return jsonify({'error': f'Datenbankfehler: {e}'}), 500
    except Exception as e:
        import traceback
        print(f"Allgemeiner Fehler in /city_chart/{city_id}: {traceback.format_exc()}")
        return jsonify({'error': f'Verarbeitungsfehler: {e}'}), 500
    finally:
        if conn:
            conn.close()

@app.route('/process', methods=['POST'])
def process():
    # Parameter aus dem Formular abrufen
    min_age = int(request.form.get('min_age', 18))
    max_age = int(request.form.get('max_age', 35))
    w_pop = float(request.form.get('w_pop', 0.3))
    w_age = float(request.form.get('w_age', 0.5))
    w_income = float(request.form.get('w_income', 0.2))
    
    # Daten verarbeiten
    try:
        result = process_data(min_age, max_age, w_pop, w_age, w_income)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/interactive_scatter', methods=['GET'])
def interactive_scatter():
    """Stellt die interaktive Scatter-Plot-Seite bereit"""
    return render_template('interactive_scatter.html')

@app.route('/interactive_clustering', methods=['GET'])
def interactive_clustering():
    """Stellt die interaktive Clustering-Seite bereit"""
    return render_template('interactive_clustering.html')

@app.route('/interactive_clustering2', methods=['GET'])
def interactive_clustering2():
    """Stellt die interaktive Clustering-Seite für Einwohner vs. Zielgruppe bereit"""
    return render_template('interactive_clustering2.html')

@app.route('/debug', methods=['GET'])
def debug_page():
    """Debug-Seite für Clustering-2"""
    return render_template('debug.html')

@app.route('/filtered_clustering', methods=['POST'])
def filtered_clustering():
    """Gibt gefilterte Clustering-Daten zurück basierend auf ausgewählten Clustern"""
    try:
        # Parameter aus der Anfrage holen
        min_age = int(request.form.get('min_age', 18))
        max_age = int(request.form.get('max_age', 35))
        w_pop = float(request.form.get('w_pop', 0.3))
        w_age = float(request.form.get('w_age', 0.5))
        w_income = float(request.form.get('w_income', 0.2))
        
        # Gewählte Cluster als Liste von Zahlen (0-4) erhalten
        selected_clusters = request.form.getlist('selected_clusters[]')
        selected_clusters = [int(cluster) for cluster in selected_clusters]
        
        # --- Daten aus DB laden (ähnlich wie in process_data) ---
        conn = None
        try:
            conn = get_db() 
            cursor = conn.cursor()
            cursor.execute("SELECT MAX(year) FROM demographics")
            latest_year = cursor.fetchone()[0]
            if not latest_year:
                raise ValueError("Keine Demografie-Daten.")

            query_main = f"""
                SELECT c.city_id, c.name AS location_name, c.simplified_name,
                       c.bundesland AS Land, d.total_population AS total,
                       d.income AS Einkommen_{latest_year}
                FROM city c JOIN demographics d ON c.city_id = d.city_id
                WHERE d.year = ?
            """
            df_main = pd.read_sql_query(query_main, conn, params=(latest_year,))

            query_age = f"""
                SELECT d.city_id, ag.label AS age_group_label, dad.count
                FROM demo_age_dist dad
                JOIN demographics d ON dad.demography_id = d.demography_id
                JOIN age_group ag ON dad.age_group_id = ag.age_group_id
                WHERE d.year = ?
            """
            df_age_dist = pd.read_sql_query(query_age, conn, params=(latest_year,))
            df_age_pivot = df_age_dist.pivot(index='city_id', columns='age_group_label', values='count').reset_index().fillna(0)
            df = pd.merge(df_main, df_age_pivot, on='city_id', how='left')
            
            if f'Einkommen_{latest_year}' in df.columns and 'Einkommen_2022' not in df.columns:
                df = df.rename(columns={f'Einkommen_{latest_year}': 'Einkommen_2022'})
            elif 'Einkommen_2022' not in df.columns:
                df['Einkommen_2022'] = 0
                
            if 'simplified_name' not in df.columns:
                 df['simplified_name'] = df['location_name']
            df['simplified_name'] = df['simplified_name'].fillna(df['location_name'])
                 
        except sqlite3.Error as e:
            print(f"DB Fehler in filtered_clustering: {e}")
            return jsonify({'error': f'Datenbankfehler: {e}'}), 500
        finally:
            if conn:
                conn.close()
        # --- Ende Datenladen aus DB ---

        # Zielgruppe berechnen
        try:
            df = calculate_target_group(df, min_age, max_age)
        except Exception as e:
            print(f"Fehler calculate_target_group in filtered_clustering: {e}")
            return jsonify({'error': f'Fehler Zielgruppenberechnung: {e}'}), 500
        
        # Normalisierung und Score-Berechnung (Code aus process_data wiederverwenden)
        for col in ['total', 'Einkommen_2022', 'target_group_percent']:
            if col not in df.columns:
                df[col] = 0
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        total_min, total_max = df["total"].min(), df["total"].max()
        income_min, income_max = df["Einkommen_2022"].min(), df["Einkommen_2022"].max()
        target_min, target_max = df["target_group_percent"].min(), df["target_group_percent"].max()
        
        df["norm_pop"] = df["total"].apply(lambda x: normalize(x, total_min, total_max) if total_max > total_min else 0)
        df["norm_income"] = df["Einkommen_2022"].apply(lambda x: normalize(x, income_min, income_max) if income_max > income_min else 0)
        df["norm_target"] = df["target_group_percent"].apply(lambda x: normalize(x, target_min, target_max) if target_max > target_min else 0)
        
        total_weight_norm = w_pop + w_age + w_income
        w_pop_norm = w_pop / total_weight_norm
        w_age_norm = w_age / total_weight_norm
        w_income_norm = w_income / total_weight_norm
        df["score"] = df["norm_pop"] * w_pop_norm + df["norm_income"] * w_income_norm + df["norm_target"] * w_age_norm
        
        # Clustering mit gefilterten Daten durchführen
        # Wichtig: generate_filtered_clustering ist in utils.py definiert
        clustering_result = generate_filtered_clustering(df, n_clusters=5, selected_clusters=selected_clusters)
        
        return jsonify(clustering_result)
    
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/filtered_clustering2', methods=['POST'])
def filtered_clustering2():
    """Gibt gefilterte Clustering-Daten für Einwohner vs. Zielgruppe zurück basierend auf ausgewählten Clustern"""
    try:
        # Parameter aus der Anfrage holen
        min_age = int(request.form.get('min_age', 18))
        max_age = int(request.form.get('max_age', 35))
        w_pop = float(request.form.get('w_pop', 0.3))
        w_age = float(request.form.get('w_age', 0.5))
        w_income = float(request.form.get('w_income', 0.2))
        
        # Gewählte Cluster als Liste von Zahlen (0-4) erhalten
        selected_clusters = request.form.getlist('selected_clusters[]')
        selected_clusters = [int(cluster) for cluster in selected_clusters]
        
        # --- Daten aus DB laden (Code wie oben) ---
        conn = None
        try:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute("SELECT MAX(year) FROM demographics")
            latest_year = cursor.fetchone()[0]
            if not latest_year:
                raise ValueError("Keine Demografie-Daten.")

            query_main = f"""
                SELECT c.city_id, c.name AS location_name, c.simplified_name,
                       c.bundesland AS Land, d.total_population AS total,
                       d.income AS Einkommen_{latest_year}
                FROM city c JOIN demographics d ON c.city_id = d.city_id
                WHERE d.year = ?
            """
            df_main = pd.read_sql_query(query_main, conn, params=(latest_year,))

            query_age = f"""
                SELECT d.city_id, ag.label AS age_group_label, dad.count
                FROM demo_age_dist dad
                JOIN demographics d ON dad.demography_id = d.demography_id
                JOIN age_group ag ON dad.age_group_id = ag.age_group_id
                WHERE d.year = ?
            """
            df_age_dist = pd.read_sql_query(query_age, conn, params=(latest_year,))
            df_age_pivot = df_age_dist.pivot(index='city_id', columns='age_group_label', values='count').reset_index().fillna(0)
            df = pd.merge(df_main, df_age_pivot, on='city_id', how='left')
            
            if f'Einkommen_{latest_year}' in df.columns and 'Einkommen_2022' not in df.columns:
                df = df.rename(columns={f'Einkommen_{latest_year}': 'Einkommen_2022'})
            elif 'Einkommen_2022' not in df.columns:
                df['Einkommen_2022'] = 0
                
            if 'simplified_name' not in df.columns:
                 df['simplified_name'] = df['location_name']
            df['simplified_name'] = df['simplified_name'].fillna(df['location_name'])

        except sqlite3.Error as e:
            print(f"DB Fehler in filtered_clustering2: {e}")
            return jsonify({'error': f'Datenbankfehler: {e}'}), 500
        finally:
            if conn:
                conn.close()
        # --- Ende Datenladen aus DB ---
        
        # Zielgruppe berechnen
        try:
            df = calculate_target_group(df, min_age, max_age)
        except Exception as e:
             print(f"Fehler calculate_target_group in filtered_clustering2: {e}")
             return jsonify({'error': f'Fehler Zielgruppenberechnung: {e}'}), 500
        
        # Score berechnen (wird intern von generate_filtered_clustering_population_target benötigt)
        for col in ['total', 'Einkommen_2022', 'target_group_percent']:
            if col not in df.columns:
                df[col] = 0
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            
        total_min, total_max = df["total"].min(), df["total"].max()
        income_min, income_max = df["Einkommen_2022"].min(), df["Einkommen_2022"].max()
        target_min, target_max = df["target_group_percent"].min(), df["target_group_percent"].max()

        df["norm_pop"] = df["total"].apply(lambda x: normalize(x, total_min, total_max) if total_max > total_min else 0)
        df["norm_income"] = df["Einkommen_2022"].apply(lambda x: normalize(x, income_min, income_max) if income_max > income_min else 0)
        df["norm_target"] = df["target_group_percent"].apply(lambda x: normalize(x, target_min, target_max) if target_max > target_min else 0)
        
        total_weight_norm = w_pop + w_age + w_income
        w_pop_norm = w_pop / total_weight_norm
        w_age_norm = w_age / total_weight_norm
        w_income_norm = w_income / total_weight_norm
        df["score"] = df["norm_pop"] * w_pop_norm + df["norm_income"] * w_income_norm + df["norm_target"] * w_age_norm
        
        # Clustering mit gefilterten Daten durchführen 
        # Diese Hilfsfunktion ist direkt hier in app.py definiert
        clustering_result = generate_filtered_clustering_population_target(df, n_clusters=5, selected_clusters=selected_clusters)
        
        return jsonify(clustering_result)
    
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


# --- Hilfsfunktion für gefiltertes Clustering 2 (bleibt in app.py) ---
def generate_filtered_clustering_population_target(df, n_clusters=5, selected_clusters=None):
    """
    Erstellt ein interaktives Clustering für Einwohner vs. Zielgruppe mit Plotly, 
    zeigt aber nur die ausgewählten Cluster an
    
    Args:
        df: DataFrame mit den Daten
        n_clusters: Anzahl der Cluster (Standard: 5)
        selected_clusters: Liste der Cluster-Indizes, die angezeigt werden sollen (0-4)
    
    Returns:
        Dictionary mit Plotly-JSON und HTML-Tabelle für die Statistiken
    """
    if selected_clusters is None:
        selected_clusters = list(range(n_clusters))  # Standardmäßig alle Cluster anzeigen
    
    # Feature-Auswahl - nur Einwohnerzahl und Zielgruppenanteil
    features = df[['total', 'target_group_percent']].copy()
    
    # Skalierung der Daten
    scaler = StandardScaler()
    scaled_features = scaler.fit_transform(features)
    
    # K-Means Clustering
    kmeans = KMeans(n_clusters=n_clusters, random_state=42)
    # Wichtig: Hier den originalen DataFrame verwenden für die Zuordnung!
    # df_clustered = df.copy() # Nicht nötig, wir fügen zu df hinzu
    
    # Cluster-Zuweisungen basierend auf K-Means
    raw_clusters = kmeans.fit_predict(scaled_features)
    
    # Zentren berechnen
    centers = scaler.inverse_transform(kmeans.cluster_centers_)
    
    # Zuordnung der Cluster zu den Interpretationen basierend auf ihren Charakteristika
    cluster_mapping = {}
    
    # Finde den Cluster mit den höchsten Bevölkerungszahlen -> Großstädte
    cluster_mapping[np.argmax([center[0] for center in centers])] = 1  # Blau
    
    # Finde den Cluster mit dem höchsten Zielgruppenanteil -> Große Universitätsstädte
    cluster_mapping[np.argmax([center[1] for center in centers])] = 3  # Lila
    
    # Finde den Cluster mit dem niedrigsten Zielgruppenanteil -> Städte mit geringem Zielgruppenanteil
    cluster_mapping[np.argmin([center[1] for center in centers])] = 2  # Grün
    
    # Finde den Cluster mit den wenigsten Einwohnern und mittlerem/hohem Zielgruppenanteil -> Kleinere Universitätsstädte
    remaining = [i for i in range(n_clusters) if i not in cluster_mapping.keys()]
    if len(remaining) >= 2:
        pop_sorted = sorted(remaining, key=lambda i: centers[i][0])
        target_sorted = sorted(pop_sorted, key=lambda i: centers[i][1], reverse=True)
        cluster_mapping[target_sorted[0]] = 4  # Orange
        for i in target_sorted[1:]:
            if i not in cluster_mapping:
                cluster_mapping[i] = 0  # Rot
    else:
        for i in remaining:
            cluster_mapping[i] = 0  # Rot
    
    # Remapping der Cluster-Zuweisungen zum DataFrame hinzufügen
    df['cluster'] = [cluster_mapping[c] for c in raw_clusters]
    
    # Formatierung für Hover-Text
    # Sicherstellen, dass Einkommen vorhanden ist
    if 'Einkommen_2022' not in df.columns:
        df['Einkommen_2022'] = 0 # Fallback
        
    df['formatted_total'] = df['total'].apply(lambda x: f"{int(x):,}".replace(",", "."))
    df['formatted_income'] = df['Einkommen_2022'].apply(lambda x: f"{int(x):,}".replace(",", "."))
    df['target_percent_display'] = (df['target_group_percent'] * 100).round(1).astype(str) + "%"
    
    # Filtere auf die ausgewählten Cluster
    df_filtered = df[df['cluster'].isin(selected_clusters)].copy()
    

    # Cluster-Beschreibungen
    cluster_descriptions = [
        'Mittelstädte',
        'Großstädte',
        'Städte mit geringem Zielgruppenanteil',
        'Große Universitätsstädte',
        'Kleinere Universitätsstädte'
    ]
    
    # Statistiken für die gefilterten Cluster berechnen
    cluster_stats = []
    for i in selected_clusters:
        cluster_data = df_filtered[df_filtered['cluster'] == i]
        if len(cluster_data) > 0:
            top_city = cluster_data.sort_values('score', ascending=False).iloc[0]['location_name']
            stats = {
                'cluster': f'Cluster {i+1}',
                'description': cluster_descriptions[i],
                'count': len(cluster_data),
                'avg_population': f"{int(cluster_data['total'].mean()):,}".replace(",", "."),
                'avg_income': f"{int(cluster_data['Einkommen_2022'].mean()):,}".replace(",", "."),
                'avg_target': (cluster_data['target_group_percent'].mean() * 100).round(1),
                'avg_score': f"{(cluster_data['score'].mean() * 100):.2f}%",
                'top_city': top_city
            }
            cluster_stats.append(stats)
        else:
            stats = {
                'cluster': f'Cluster {i+1}',
                'description': cluster_descriptions[i],
                'count': 0,
                'avg_population': "0",
                'avg_income': "0",
                'avg_target': 0.0,
                'avg_score': '0,00%',
                'top_city': 'N/A'
            }
            cluster_stats.append(stats)
    
    # Definiere Clusterfarben
    cluster_colors = ['#e41a1c', '#377eb8', '#4daf4a', '#984ea3', '#ff7f00']
    
    # Cluster-Namen für die Legende
    cluster_names = {
        0: f'Cluster 1: {cluster_descriptions[0]}',
        1: f'Cluster 2: {cluster_descriptions[1]}',
        2: f'Cluster 3: {cluster_descriptions[2]}',
        3: f'Cluster 4: {cluster_descriptions[3]}',
        4: f'Cluster 5: {cluster_descriptions[4]}'
    }
    
    # DataFrame für Cluster-Mapping vorbereiten
    df_filtered['cluster_name'] = df_filtered['cluster'].map(cluster_names)
    
    # Erstelle Plotly-Figur
    fig = px.scatter(
        df_filtered,  # Nur gefilterte Daten anzeigen
        x='total',
        y='target_group_percent',
        color='cluster_name',  # Verwende die benannten Cluster
        hover_name='location_name',
        hover_data={
            'total': False,
            'target_group_percent': False,
            'cluster_name': True,
            'formatted_total': True,
            'formatted_income': True,
            'target_percent_display': True,
            'Land': True,
            'score': True
        },
        labels={
            'formatted_total': 'Einwohner',
            'formatted_income': 'Einkommen',
            'target_percent_display': 'Zielgruppe',
            'Land': 'Bundesland',
            'score': 'Score',
            'cluster_name': 'Cluster'
        },
        color_discrete_map={
            f'Cluster 1: {cluster_descriptions[0]}': cluster_colors[0],
            f'Cluster 2: {cluster_descriptions[1]}': cluster_colors[1],
            f'Cluster 3: {cluster_descriptions[2]}': cluster_colors[2],
            f'Cluster 4: {cluster_descriptions[3]}': cluster_colors[3],
            f'Cluster 5: {cluster_descriptions[4]}': cluster_colors[4]
        },
        size='norm_target',
        size_max=20,
        template='plotly_white',
        category_orders={'cluster_name': [
            f'Cluster 1: {cluster_descriptions[0]}',
            f'Cluster 2: {cluster_descriptions[1]}', 
            f'Cluster 3: {cluster_descriptions[2]}',
            f'Cluster 4: {cluster_descriptions[3]}', 
            f'Cluster 5: {cluster_descriptions[4]}'
        ]}
    )
    
    # Linien um Marker anpassen
    fig.update_traces(
        marker=dict(
            line=dict(width=1, color='DarkSlateGrey')
        )
    )
    
    # Layout anpassen
    fig.update_layout(
        title='Clustering der Städte nach Einwohnerzahl und Zielgruppenanteil',
        xaxis_title='Einwohnerzahl',
        yaxis_title='Zielgruppenanteil',
        height=600,
        legend_title='Cluster',
        hovermode='closest'
    )
    
    # Clusterzentren hinzufügen (nur für aktive Cluster)
    # Aber Clusterzentren auch für gefilterte Cluster anzeigen, zur Einordnung
    # Verwende die gemappten Cluster-IDs (0-4)
    centers_df = pd.DataFrame({
        'total': [centers[i][0] for i in range(n_clusters)],
        'target_group_percent': [centers[i][1] for i in range(n_clusters)],
        'cluster_name': [f'Zentrum C{cluster_mapping[i]+1}' for i in range(n_clusters)] 
    })
    # Filtere Zentren basierend auf den ausgewählten Clustern
    centers_df_filtered = centers_df[centers_df['cluster_name'].str.contains(f'C[{ "".join(map(str, [c+1 for c in selected_clusters])) }]')]

    
    # Zentren als Kreuze darstellen (nur wenn welche übrig sind)
    if not centers_df_filtered.empty:
        fig.add_trace(
            px.scatter(
                centers_df_filtered, 
                x='total', 
                y='target_group_percent', 
                text='cluster_name',
                color_discrete_sequence=['black']
            ).update_traces(
                marker=dict(symbol='x', size=15, line=dict(width=2)),
                mode='markers+text',
                textposition='top center'
            ).data[0]
        )
    
    # Zusammenfassung der Cluster-Statistiken als HTML-Tabelle
    stats_df = pd.DataFrame(cluster_stats)
    stats_df.columns = [
        'Cluster', 'Beschreibung', 'Anzahl Städte', 'Durchschn. Einwohner', 'Durchschn. Einkommen',
        'Durchschn. Zielgruppe (%)', 'Durchschn. Score', 'Top Stadt'
    ]
    
    # Als JSON und HTML zurückgeben
    return {
        'plot': fig.to_json(),
        'stats': stats_df.to_html(
            classes='table table-striped table-hover',
            index=False,
            justify='left',
            border=0
        )
    }

@app.route('/export_selected', methods=['POST'])
def export_selected():
    """Exportiert die ausgewählten Städte mit den aktuellen Parametern als Excel."""
    try:
        request_data = request.get_json()
        selected_city_ids_str = request_data.get('selected_cities', [])
        # Stelle sicher, dass IDs Integer sind
        try:
            selected_city_ids = [int(sid) for sid in selected_city_ids_str]
        except (ValueError, TypeError) as e:
             print(f"Fehler beim Konvertieren der City IDs: {e}, Daten: {selected_city_ids_str}")
             return jsonify({'error': 'Ungültige Stadt-IDs übermittelt.'}), 400

        if not selected_city_ids:
            return jsonify({'error': 'Keine Städte für den Export ausgewählt.'}), 400

        # Parameter aus dem Request holen (mit Fallbacks)
        min_age = int(request_data.get('min_age', 18))
        max_age = int(request_data.get('max_age', 35))
        # Beachte: JS sendet 0.0-1.0, nicht 0-100
        w_pop = float(request_data.get('w_pop', 0.3))
        w_age = float(request_data.get('w_age', 0.5))
        w_income = float(request_data.get('w_income', 0.2))

        print(f"Export angefordert für {len(selected_city_ids)} Städte mit Parametern: Age({min_age}-{max_age}), Weights(P:{w_pop}, A:{w_age}, I:{w_income})")

        # === WICHTIG: Datenverarbeitung mit den aktuellen Parametern WIEDERHOLEN ===
        # Hole die *kompletten*, *aktuell berechneten* Daten für ALLE Städte
        # Wir verwenden die `process_data`-Funktion (oder Teile davon) erneut
        # Achtung: process_data gibt HTML zurück, wir brauchen das DataFrame!
        # => Wir müssen die Daten-Generierung und -Berechnung von process_data extrahieren
        # oder process_data anpassen, damit es optional das DataFrame zurückgibt.
        
        # === Extraktion der Logik aus process_data (vereinfacht) ===
        conn = None
        try:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute("SELECT MAX(year) FROM demographics")
            latest_year = cursor.fetchone()[0]
            if not latest_year: raise ValueError("Keine Demografie-Daten.")

            query_main = f"""
                WITH CityEventCounts AS (
                    -- Zähle Vorkommen jedes Event-Textes pro Stadt
                    SELECT
                        pc.city_id,
                        peg.event_gastro_text,
                        COUNT(peg.event_gastro_text) AS text_count
                    FROM postal_code pc
                    JOIN plz_event_gastro peg ON pc.postal_code_id = peg.postal_code_id
                    WHERE peg.event_gastro_text IS NOT NULL AND peg.event_gastro_text != ''
                    GROUP BY pc.city_id, peg.event_gastro_text
                ),
                RankedCityEvents AS (
                    -- Weise jedem Text pro Stadt einen Rang basierend auf der Häufigkeit zu
                    SELECT
                        city_id,
                        event_gastro_text,
                        ROW_NUMBER() OVER (PARTITION BY city_id ORDER BY text_count DESC, event_gastro_text ASC) AS rn
                    FROM CityEventCounts
                )
                SELECT
                    c.city_id, c.name AS location_name, c.name, c.region_code, -- c.name doppelt auswählen
                    c.latitude, c.longitude, c.simplified_name, c.bundesland,
                    d.total_population AS total,
                    d.income AS Einkommen_{latest_year},
                    -- Hole den Text mit Rang 1 (den häufigsten) oder 'N/A'
                    COALESCE(rce.event_gastro_text, 'N/A') AS event_gastro_text
                FROM city c
                JOIN demographics d ON c.city_id = d.city_id
                LEFT JOIN RankedCityEvents rce ON c.city_id = rce.city_id AND rce.rn = 1 -- Join hinzugefügt
                WHERE d.year = ?
            """
            df_main = pd.read_sql_query(query_main, conn, params=(latest_year,))

            query_age = f""" 
                SELECT d.city_id, ag.label AS age_group_label, dad.count
                FROM demo_age_dist dad 
                JOIN demographics d ON dad.demography_id = d.demography_id 
                JOIN age_group ag ON dad.age_group_id = ag.age_group_id WHERE d.year = ?
            """
            df_age_dist = pd.read_sql_query(query_age, conn, params=(latest_year,))
            df_age_pivot = df_age_dist.pivot(index='city_id', columns='age_group_label', values='count').reset_index().fillna(0)
            df = pd.merge(df_main, df_age_pivot, on='city_id', how='left')
            
            if f'Einkommen_{latest_year}' in df.columns and 'Einkommen_2022' not in df.columns:
                 df = df.rename(columns={f'Einkommen_{latest_year}': 'Einkommen_2022'})
            elif 'Einkommen_2022' not in df.columns: df['Einkommen_2022'] = 0
            if 'simplified_name' not in df.columns: df['simplified_name'] = df['name']
            df['simplified_name'] = df['simplified_name'].fillna(df['name'])
            if 'bundesland' not in df.columns: df['bundesland'] = 'N/A' # Fallback
            df['bundesland'] = df['bundesland'].fillna('N/A')

            # Zielgruppe berechnen
            df = calculate_target_group(df, min_age, max_age)
            
            # Normalisierung und Score-Berechnung
            for col in ['total', 'Einkommen_2022', 'target_group_percent']:
                if col not in df.columns: df[col] = 0
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

            total_min, total_max = df["total"].min(), df["total"].max()
            income_min, income_max = df["Einkommen_2022"].min(), df["Einkommen_2022"].max()
            target_min, target_max = df["target_group_percent"].min(), df["target_group_percent"].max()
            df["norm_pop"] = df["total"].apply(lambda x: normalize(x, total_min, total_max) if total_max > total_min else 0)
            df["norm_income"] = df["Einkommen_2022"].apply(lambda x: normalize(x, income_min, income_max) if income_max > income_min else 0)
            df["norm_target"] = df["target_group_percent"].apply(lambda x: normalize(x, target_min, target_max) if target_max > target_min else 0)
            
            total_weight_norm = w_pop + w_age + w_income
            if total_weight_norm == 0: total_weight_norm = 1 # Verhindere Division durch 0
            w_pop_norm = w_pop / total_weight_norm
            w_age_norm = w_age / total_weight_norm
            w_income_norm = w_income / total_weight_norm
            df["score"] = df["norm_pop"] * w_pop_norm + df["norm_income"] * w_income_norm + df["norm_target"] * w_age_norm
            
            # Clustering durchführen (vereinfacht, nur zur Namensfindung)
            features_for_clustering = df[['total', 'Einkommen_2022', 'target_group_percent']].fillna(0)
            scaler = StandardScaler()
            scaled_features = scaler.fit_transform(features_for_clustering)
            kmeans = KMeans(n_clusters=5, random_state=42)
            raw_clusters = kmeans.fit_predict(scaled_features)
            centers = scaler.inverse_transform(kmeans.cluster_centers_)
            cluster_mapping = {}
            cluster_mapping[np.argmax([center[0] for center in centers])] = 1
            cluster_mapping[np.argmax([center[2] for center in centers])] = 3
            cluster_mapping[np.argmax([center[1] for center in centers])] = 4
            cluster_mapping[np.argmin([center[1] for center in centers])] = 2
            for i in range(5): 
                if i not in cluster_mapping: cluster_mapping[i] = 0
            cluster_names = {0: "Mittelstädte", 1: "Großstädte", 2: "Ländliche Regionen", 3: "Universitätsstädte", 4: "Wohlhabende Mittelstädte"}
            df['cluster_id'] = [cluster_mapping[c] for c in raw_clusters]
            df['cluster_name'] = df['cluster_id'].map(cluster_names)
            
        except sqlite3.Error as e:
            print(f"DB Fehler in /export_selected: {e}")
            return jsonify({'error': f'Datenbankfehler: {e}'}), 500
        except Exception as e:
             import traceback
             print(f"Allgemeiner Fehler in /export_selected Datenaufbereitung: {traceback.format_exc()}")
             return jsonify({'error': f'Fehler bei Datenaufbereitung: {e}'}), 500
        finally:
            if conn: conn.close()
        # === Ende Datenaufbereitung ===


        # Filtere das DataFrame auf die ausgewählten Städte
        df_selected = df[df['city_id'].isin(selected_city_ids)].copy()

        if df_selected.empty:
             return jsonify({'error': 'Keine der ausgewählten Städte in den verarbeiteten Daten gefunden.'}), 404

        # Sortiere nach Score (absteigend) und füge Platz hinzu
        df_selected = df_selected.sort_values("score", ascending=False).reset_index(drop=True)
        df_selected.index = df_selected.index + 1
        df_selected = df_selected.reset_index().rename(columns={'index': 'Platz'})

        # Wähle und benenne Spalten für den Export
        export_columns = {
            'Platz': 'Platz',
            'name': 'Stadt', # Den ursprünglichen Namen verwenden
            'bundesland': 'Bundesland',
            'total': 'Einwohner',
            'Einkommen_2022': 'Einkommen',
            'target_group_percent': 'Zielgruppe',
            'cluster_name': 'Cluster',
            'score': 'Score',
            'event_gastro_text': 'Vertriebsnummer' # Hinzugefügt
        }
        df_export = df_selected[list(export_columns.keys())].rename(columns=export_columns)

        # Formatiere die Spalten für Excel
        df_export['Zielgruppe'] = (df_export['Zielgruppe'] * 100).round(1) # Als Zahl belassen für Excel
        df_export['Score'] = (df_export['Score'] * 100).round(2) # Als Zahl belassen für Excel
        df_export['Einwohner'] = df_export['Einwohner'].astype(int)
        df_export['Einkommen'] = df_export['Einkommen'].astype(int)

        # Erstelle Excel-Datei im Speicher
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Ausgewählte Städte')
            
            # Zugriff auf das Worksheet für Formatierungen
            workbook = writer.book
            worksheet = writer.sheets['Ausgewählte Städte']
            
            # Formatierungen anwenden
            header_font = Font(bold=True)
            centered_alignment = Alignment(horizontal='center', vertical='center')
            right_alignment = Alignment(horizontal='right', vertical='center')
            percent_format_1_decimal = '0.0"%"' # Format für % mit einer Dezimalstelle
            percent_format_2_decimals = '0.00"%"' # Format für % mit zwei Dezimalstellen
            number_format_thousands = '#,##0' # Format für Zahlen mit Tausenderpunkt

            for col_idx, column_name in enumerate(df_export.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.alignment = centered_alignment
                
                column_letter = chr(64 + col_idx)
                
                # Spaltenbreiten anpassen (geschätzt)
                if column_name == 'Stadt':
                    worksheet.column_dimensions[column_letter].width = 30
                elif column_name == 'Bundesland':
                     worksheet.column_dimensions[column_letter].width = 20
                elif column_name == 'Cluster':
                     worksheet.column_dimensions[column_letter].width = 25
                elif column_name in ['Einwohner', 'Einkommen']:
                     worksheet.column_dimensions[column_letter].width = 15
                     # Zahlenformat anwenden
                     for row_idx in range(2, worksheet.max_row + 1):
                         worksheet[f"{column_letter}{row_idx}"].number_format = number_format_thousands
                         worksheet[f"{column_letter}{row_idx}"].alignment = right_alignment
                elif column_name == 'Zielgruppe':
                     worksheet.column_dimensions[column_letter].width = 12
                     # Prozentformat anwenden
                     for row_idx in range(2, worksheet.max_row + 1):
                         worksheet[f"{column_letter}{row_idx}"].number_format = percent_format_1_decimal
                         worksheet[f"{column_letter}{row_idx}"].alignment = right_alignment
                elif column_name == 'Score':
                     worksheet.column_dimensions[column_letter].width = 10
                     # Prozentformat anwenden
                     for row_idx in range(2, worksheet.max_row + 1):
                         worksheet[f"{column_letter}{row_idx}"].number_format = percent_format_2_decimals
                         worksheet[f"{column_letter}{row_idx}"].alignment = right_alignment
                elif column_name == 'Platz':
                     worksheet.column_dimensions[column_letter].width = 8
                     for row_idx in range(2, worksheet.max_row + 1):
                         worksheet[f"{column_letter}{row_idx}"].alignment = centered_alignment
                elif column_name == 'Vertriebsnummer': # Formatierung für neue Spalte hinzugefügt
                     worksheet.column_dimensions[column_letter].width = 20
                     # Standardmäßig linksbündig, keine spezielle Zahlenformatierung
                else:
                    worksheet.column_dimensions[column_letter].width = 15


        output.seek(0)

        # Aktuelles Datum für Dateinamen
        today = datetime.now().strftime("%Y%m%d")
        filename = f"Stadtanalyse_Export_{today}.xlsx"

        # Datei zum Download senden
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        print(f"Fehler in /export_selected: {traceback.format_exc()}")
        return jsonify({'error': f'Interner Serverfehler beim Export: {e}'}), 500

@app.route('/live_search/<stadt_search_name>/<suchbegriff>', methods=['POST']) # Parameter umbenannt für Klarheit
def live_search(stadt_search_name, suchbegriff):
    """Führt eine vereinfachte Live-Suche durch, speichert das Ergebnis und gibt es zurück."""
    if not API_KEY or API_KEY == 'DEIN_API_KEY':
         return jsonify({'error': 'Google Maps API Key nicht konfiguriert.'}), 500

    # Hole simplified_name und Koordinaten aus der Datenbank
    city_latitude = None
    city_longitude = None
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            "SELECT simplified_name, latitude, longitude FROM city WHERE simplified_name = ? OR name = ?",
            (stadt_search_name, stadt_search_name)
        )
        city_data = cursor.fetchone()
        db.close()

        if city_data and city_data['simplified_name']:
            city_simplified_name = city_data['simplified_name']
            city_latitude = city_data.get('latitude')
            city_longitude = city_data.get('longitude')
        else:
            # Fallback: verwende den übergebenen Namen
            city_simplified_name = stadt_search_name
            print(f"Warnung: Kein simplified_name für '{stadt_search_name}' gefunden, verwende Original-Namen")

    except Exception as e:
        print(f"Fehler beim Laden des simplified_name für {stadt_search_name}: {e}")
        city_simplified_name = stadt_search_name  # Fallback

    # Erstelle die Suchanfrage
    query = f"{suchbegriff} in {city_simplified_name}"
    print(f"Starte Live-Suche für: {query}")
    if city_latitude and city_longitude:
        print(f"Mit Koordinaten: lat={city_latitude}, lon={city_longitude}")

    # --- API Aufruf ---
    search_result_api = fetch_google_maps_data(API_KEY, query, latitude=city_latitude, longitude=city_longitude)

    if "error" in search_result_api:
        status_code = search_result_api.get('status_code', 500)
        return jsonify({'error': search_result_api.get('error', 'Unbekannter API Fehler'), 'data': {'places': search_result_api.get('places', [])} }), status_code

    places_from_api = search_result_api.get("places", [])
    
    # Explizit loggen, wenn keine Ergebnisse gefunden wurden
    if len(places_from_api) == 0:
        print(f"Live-Suche für '{query}': Keine Ergebnisse gefunden (0 Orte).")
    else:
        print(f"Live-Suche für '{query}': {len(places_from_api)} Orte gefunden.")

    # Sortieren nach Rating (absteigend), sekundär nach userRatingCount (absteigend)
    sorted_places = sorted(
        places_from_api,
        key=lambda p: (p.get('rating', 0), p.get('userRatingCount', 0)),
        reverse=True
    )
    # Nur die ersten 20 anzeigen (sollte eigentlich schon von der API begrenzt sein)
    final_places_to_show = sorted_places[:20]

    # --- Ergebnis speichern (ohne Filter-Parameter) ---
    try:
        response_json_str = json.dumps({"places": places_from_api})  # Speichere alle Ergebnisse
        now = datetime.now().isoformat()
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            """
            INSERT INTO search_cache (stadt, suchbegriff, response_json, last_updated)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(stadt, suchbegriff) DO UPDATE SET
                response_json = excluded.response_json,
                last_updated = excluded.last_updated
            """,
            (stadt_search_name, suchbegriff, response_json_str, now)
        )
        db.commit()
        db.close()
        print(f"Ergebnis für '{query}' erfolgreich gespeichert/aktualisiert.")
        return jsonify({ 
            'last_updated': datetime.fromisoformat(now).strftime('%d.%m.%Y %H:%M:%S'),
            'data': {"places": final_places_to_show} 
        })
    except Exception as e:
        print(f"Fehler beim Speichern des Ergebnisses in DB: {e}")
        return jsonify({ 
            'data': {"places": final_places_to_show}, 
            'warning': f'Ergebnis konnte nicht gespeichert werden: {e}'
        }), 200

@app.route('/get_search_info/<stadt_search_name>/<suchbegriff>', methods=['GET']) # Parameter umbenannt
def get_search_info(stadt_search_name, suchbegriff):
    """Prüft, ob und wann die letzte Suche für eine Stadt/Begriff stattfand."""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            "SELECT last_updated FROM search_cache WHERE stadt = ? AND suchbegriff = ? ORDER BY last_updated DESC LIMIT 1",
            (stadt_search_name, suchbegriff)
        )
        result = cursor.fetchone()
        db.close()

        if result:
            last_updated_str = datetime.fromisoformat(result['last_updated']).strftime('%d.%m.%Y %H:%M:%S')
            return jsonify({
                'last_updated': last_updated_str,
                'min_rating': result['min_rating'], # Kann None sein
                'min_user_ratings': result['min_user_ratings'] # Kann None sein (wird als 0 behandelt)
            })
        else:
            return jsonify({'last_updated': None})
    except Exception as e:
        print(f"Fehler beim Abrufen der Suchinfo: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/get_cached_data/<stadt_search_name>/<suchbegriff>', methods=['GET']) # Parameter umbenannt
def get_cached_data(stadt_search_name, suchbegriff):
    """Holt die gespeicherten Suchergebnisse aus der Datenbank."""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            "SELECT response_json, last_updated FROM search_cache WHERE stadt = ? AND suchbegriff = ? ORDER BY last_updated DESC LIMIT 1",
            (stadt_search_name, suchbegriff)
        )
        result = cursor.fetchone()
        db.close()

        if result and result['response_json']:
            try:
                 data_from_db = json.loads(result['response_json'])
                 places_from_db_cache = data_from_db.get("places", [])
                 
                 # Sortiere die Daten aus dem Cache
                 sorted_places_cache = sorted(
                     places_from_db_cache,
                     key=lambda p: (p.get('rating', 0), p.get('userRatingCount', 0)),
                     reverse=True
                 )
                 final_places_to_show_cache = sorted_places_cache[:20]

                 last_updated_str = datetime.fromisoformat(result['last_updated']).strftime('%d.%m.%Y %H:%M:%S')
                 return jsonify({
                    'last_updated': last_updated_str,
                    'data': {"places": final_places_to_show_cache} 
                 })
            except json.JSONDecodeError:
                 return jsonify({'error': 'Gespeicherte Daten sind ungültig (JSON Fehler)'}), 500
        else:
            return jsonify({'error': 'Keine gespeicherten Daten gefunden'}), 404
    except Exception as e:
        print(f"Fehler beim Abrufen der Cache-Daten: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/start_search/<term_name>', methods=['POST'])
def start_search(term_name):
    """Startet die Massensuche für einen gegebenen Suchbegriff im Hintergrund."""
    global API_KEY
    if not API_KEY or API_KEY == 'DEIN_API_KEY':
         return jsonify({'error': 'Google Maps API Key nicht konfiguriert.'}), 500

    # Starte die Suche in einem separaten Thread, um den Request nicht zu blockieren
    import threading
    thread = threading.Thread(
        target=run_place_search_for_all_cities, 
        args=(term_name, API_KEY)
    )
    thread.daemon = True # Thread stirbt, wenn Hauptprogramm endet
    thread.start()
    
    return jsonify({'message': f'Suche für "{term_name}" gestartet.'}), 202 # Accepted

@app.route('/search_status')
def search_status():
    """Streamt Statusmeldungen der laufenden Suche mittels SSE."""
    def generate():
        global search_status_queue
        retry_timeout = 2000  # Retry-Timeout in ms für den Client
        
        try:
            # Sende einen initialen keep-alive, damit die Verbindung etabliert wird
            yield f"retry: {retry_timeout}\n"
            yield f"data: Verbindung hergestellt. Warte auf Statusmeldungen...\n\n"
            
            while True:
                try:
                    # Warte auf eine neue Nachricht (mit Timeout, um blockieren zu verhindern)
                    message = search_status_queue.get(timeout=5)  # Reduziert auf 5 Sekunden für schnelleres Feedback
                    
                    if message == "DONE":
                        # Sende eindeutige Abschlussnachricht und beende
                        yield f"data: DONE\n\n"
                        # Sende noch eine explizite Abschlussmeldung
                        yield f"data: Suche abgeschlossen. Die Verbindung wird jetzt geschlossen.\n\n"
                        break  # Beende den Stream
                    elif message.startswith("FEHLER"):
                        # Sende Fehlermeldung als 'error' Event
                        yield f"event: error\ndata: {message}\n\n"
                    else:
                        # Sende normale Statusmeldung
                        yield f"data: {message}\n\n"
                    
                    search_status_queue.task_done()
                    
                except queue.Empty:
                    # Sende keep-alive mit retry-Anweisung
                    yield f"retry: {retry_timeout}\n"
                    yield f": keep-alive\n\n"
        
        except GeneratorExit:
            # Client hat die Verbindung geschlossen
            print("SSE-Verbindung vom Client getrennt")
            return
        except Exception as e:
            error_msg = f"Fehler im SSE Generator: {e}"
            print(error_msg)
            yield f"event: error\ndata: {error_msg}\n\n"
            yield f"data: Verbindung wird aufgrund eines Fehlers beendet.\n\n"

    # Server-Sent Event Header setzen
    response = Response(stream_with_context(generate()), 
                        mimetype='text/event-stream')
    
    # CORS-Header für SSE
    response.headers['Cache-Control'] = 'no-cache'
    response.headers['X-Accel-Buffering'] = 'no'  # Für Nginx
    response.headers['Connection'] = 'keep-alive'
    
    return response

@app.route('/get_search_terms', methods=['GET'])
def get_search_terms():
    """Gibt alle verfügbaren Suchbegriffe aus der Datenbank zurück."""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute("SELECT term_id, name FROM search_term ORDER BY name")
        terms = cursor.fetchall()
        db.close()
        # Konvertiere Rows in Dictionaries
        terms_list = [{'id': row['term_id'], 'name': row['name']} for row in terms]
        return jsonify(terms_list)
    except Exception as e:
        print(f"Fehler beim Abrufen der Suchbegriffe: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/add_search_term', methods=['POST'])
def add_search_term():
    """Fügt einen neuen Suchbegriff zur Datenbank hinzu."""
    term_name = request.json.get('term_name')
    if not term_name or not isinstance(term_name, str) or len(term_name.strip()) == 0:
        return jsonify({'error': 'Ungültiger Suchbegriff übermittelt.'}), 400

    term_name = term_name.strip()

    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute("INSERT INTO search_term (name) VALUES (?)", (term_name,))
        new_term_id = cursor.lastrowid
        db.commit()
        db.close()
        return jsonify({'message': f'Suchbegriff "{term_name}" hinzugefügt.', 'new_term': {'id': new_term_id, 'name': term_name}}), 201
    except sqlite3.IntegrityError:
        # Begriff existiert bereits
        db.close()
        return jsonify({'error': f'Suchbegriff "{term_name}" existiert bereits.'}), 409
    except Exception as e:
        db.close()
        print(f"Fehler beim Hinzufügen des Suchbegriffs '{term_name}': {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/get_keyword_results_for_cities', methods=['POST'])
def get_keyword_results_for_cities():
    """Ruft Ergebnisse von Keyword-Suchen für ausgewählte Städte ab."""
    data = request.json
    city_ids = data.get('city_ids', [])
    keyword_ids = data.get('keyword_ids', [])
    
    if not city_ids:
        return jsonify({'error': 'Keine Städte ausgewählt'}), 400
    
    if not keyword_ids:
        return jsonify({'error': 'Keine Keywords ausgewählt'}), 400
    
    try:
        db = get_db()
        cursor = db.cursor()
        
        results = {}
        
        # Für jede Stadt
        for city_id in city_ids:
            # Hole Stadt-Namen
            cursor.execute("SELECT name FROM city WHERE city_id = ?", (city_id,))
            city_result = cursor.fetchone()
            
            if not city_result:
                continue
                
            city_name = city_result['name']
            results[city_id] = {
                'city_name': city_name,
                'places': []
            }
            
            # Hole alle Orte, die in der Stadt durch eine der ausgewählten Suchbegriffe gefunden wurden
            place_ids_query = """
                SELECT DISTINCT place_id FROM place_search 
                WHERE city_id = ? AND term_id IN ({})
            """.format(','.join(['?'] * len(keyword_ids)))
            
            cursor.execute(place_ids_query, [city_id] + keyword_ids)
            place_id_results = cursor.fetchall()
            
            if not place_id_results:
                # Explizit loggen, wenn keine Ergebnisse für diese Stadt gefunden wurden
                print(f"Keine Ergebnisse für Stadt '{city_name}' (ID: {city_id}) mit den ausgewählten Keywords gefunden.")
                # Leere places-Liste beibehalten, damit die Stadt trotzdem in den Ergebnissen erscheint
                continue
                
            place_ids = [r['place_id'] for r in place_id_results]
            
            # Hole alle Details zu den gefundenen Orten, inklusive Öffnungszeiten und Summary
            places_query = f"""
                SELECT 
                    p.*, 
                    oh.weekday_text as opening_hours_text
                FROM place p 
                LEFT JOIN opening_hours oh ON p.place_id = oh.place_id
                WHERE p.place_id IN ({','.join(['?'] * len(place_ids))})
            """
            
            cursor.execute(places_query, place_ids)
            places = cursor.fetchall()
            
            # Konvertiere Row-Objekte in Dictionaries
            places_list = []
            for place in places:
                place_dict = dict(place)
                
                # Hole die neueste Bewertung für diesen Ort (Gesamtbewertung)
                cursor.execute("""
                    SELECT rating, user_rating_count FROM rating_history 
                    WHERE place_id = ? 
                    ORDER BY timestamp DESC LIMIT 1
                """, (place['place_id'],))
                
                rating_data = cursor.fetchone()
                if rating_data:
                    place_dict['rating'] = rating_data['rating']
                    place_dict['user_rating_count'] = rating_data['user_rating_count']
                
                # Hole die Keywords, mit denen dieser Ort gefunden wurde
                cursor.execute("""
                    SELECT t.name FROM search_term t
                    JOIN place_search ps ON t.term_id = ps.term_id
                    WHERE ps.place_id = ? AND ps.city_id = ? AND t.term_id IN ({})
                """.format(','.join(['?'] * len(keyword_ids))), 
                [place['place_id'], city_id] + keyword_ids)
                
                found_keywords = cursor.fetchall()
                place_dict['keywords'] = [kw['name'] for kw in found_keywords]
                
                # Für die Anzeige: Stelle sicher, dass alle Places eine displayName-Struktur haben
                # Verwende den display_name aus der Datenbank falls vorhanden, sonst den normalen Namen
                if place_dict.get('display_name'):
                    place_dict['displayName'] = {'text': place_dict['display_name']}
                elif 'displayName' not in place_dict or not place_dict['displayName']:
                    place_dict['displayName'] = {'text': place_dict['name']}
                
                # Hole alle Reviews für diesen Ort (neueste zuerst)
                cursor.execute("""
                    SELECT 
                        author_name, rating, relative_publish_time_description, text
                    FROM review
                    WHERE place_id = ?
                    ORDER BY publish_time DESC
                    -- LIMIT 5 -- Entfernt, um alle Reviews anzuzeigen
                """, (place['place_id'],))
                reviews_data = cursor.fetchall()
                place_dict['reviews'] = [dict(rev) for rev in reviews_data] # Alle Reviews
                
                places_list.append(place_dict)
            
            # Sortiere die Ergebnisliste nach Bewertung (falls vorhanden), dann nach Anzahl Bewertungen
            places_list.sort(key=lambda x: (
                x.get('rating') if x.get('rating') is not None else -1, 
                x.get('user_rating_count') if x.get('user_rating_count') is not None else -1
            ), reverse=True)
            
            results[city_id]['places'] = places_list
        
        db.close()
        return jsonify({'results': results})
    
    except Exception as e:
        print(f"Fehler beim Abrufen der Keyword-Ergebnisse: {e}")
        if db:
            db.close()
        return jsonify({'error': f'Datenbankfehler: {str(e)}'}), 500

@app.route('/keyword_search_results')
def keyword_search_results():
    """Rendert die dedizierte Seite für Keyword-Suchergebnisse."""
    city_ids_str = request.args.get('cities', '')
    keyword_ids_str = request.args.get('keywords', '')

    if not city_ids_str or not keyword_ids_str:
        return render_template('keyword_search_results.html', results=None, error="Keine Städte oder Keywords für die Suche angegeben.")

    try:
        city_ids = [int(cid) for cid in city_ids_str.split(',')]
        keyword_ids = [int(kid) for kid in keyword_ids_str.split(',')]
    except ValueError:
        return render_template('keyword_search_results.html', results=None, error="Ungültige Städte- oder Keyword-IDs.")

    # --- Daten aus der Datenbank holen (Logik aus /get_keyword_results_for_cities wiederverwendet) ---
    db = None
    try:
        db = get_db()
        cursor = db.cursor()
        results = {}

        for city_id in city_ids:
            cursor.execute("SELECT name FROM city WHERE city_id = ?", (city_id,))
            city_result = cursor.fetchone()
            if not city_result: continue
            city_name = city_result['name']
            results[city_id] = {'city_name': city_name, 'places': []}

            place_ids_query = """
                SELECT DISTINCT place_id FROM place_search 
                WHERE city_id = ? AND term_id IN ({})
            """.format(','.join(['?'] * len(keyword_ids)))
            cursor.execute(place_ids_query, [city_id] + keyword_ids)
            place_id_results = cursor.fetchall()
            
            if not place_id_results:
                # Explizit loggen, wenn keine Ergebnisse für diese Stadt gefunden wurden
                print(f"Keine Ergebnisse für Stadt '{city_name}' (ID: {city_id}) mit den ausgewählten Keywords gefunden.")
                # Leere places-Liste beibehalten, damit die Stadt trotzdem in den Ergebnissen erscheint
                continue
                
            place_ids = [r['place_id'] for r in place_id_results]

            places_query = f"""
                SELECT 
                    p.*, 
                    oh.weekday_text as opening_hours_text
                FROM place p 
                LEFT JOIN opening_hours oh ON p.place_id = oh.place_id
                WHERE p.place_id IN ({','.join(['?'] * len(place_ids))})
            """
            cursor.execute(places_query, place_ids)
            places = cursor.fetchall()

            places_list = []
            for place in places:
                place_dict = dict(place)
                cursor.execute("""
                    SELECT rating, user_rating_count FROM rating_history 
                    WHERE place_id = ? 
                    ORDER BY timestamp DESC LIMIT 1
                """, (place['place_id'],))
                rating_data = cursor.fetchone()
                if rating_data:
                    place_dict['rating'] = rating_data['rating']
                    place_dict['user_rating_count'] = rating_data['user_rating_count']
                
                cursor.execute("""
                    SELECT t.name FROM search_term t
                    JOIN place_search ps ON t.term_id = ps.term_id
                    WHERE ps.place_id = ? AND ps.city_id = ? AND t.term_id IN ({})
                """.format(','.join(['?'] * len(keyword_ids))), 
                [place['place_id'], city_id] + keyword_ids)
                found_keywords = cursor.fetchall()
                place_dict['keywords'] = [kw['name'] for kw in found_keywords]

                if place_dict.get('display_name'):
                    place_dict['displayName'] = {'text': place_dict['display_name']}
                elif 'displayName' not in place_dict or not place_dict['displayName']:
                    place_dict['displayName'] = {'text': place_dict['name']}

                cursor.execute("""
                    SELECT 
                        author_name, rating, relative_publish_time_description, text
                    FROM review
                    WHERE place_id = ?
                    ORDER BY publish_time DESC
                    -- LIMIT 5 -- Entfernt, um alle Reviews anzuzeigen
                """, (place['place_id'],))
                reviews_data = cursor.fetchall()
                place_dict['reviews'] = [dict(rev) for rev in reviews_data] # Alle Reviews
                
                places_list.append(place_dict)
            
            # Sortiere die Ergebnisliste nach Bewertung (falls vorhanden), dann nach Anzahl Bewertungen
            places_list.sort(key=lambda x: (
                x.get('rating') if x.get('rating') is not None else -1, 
                x.get('user_rating_count') if x.get('user_rating_count') is not None else -1
            ), reverse=True)
            
            results[city_id]['places'] = places_list
        
        db.close()
        return render_template('keyword_search_results.html', results=results)

    except sqlite3.Error as db_err:
        error_msg = f'Datenbankfehler: {str(db_err)}'
        print(f"Fehler in /keyword_search_results: {error_msg}")
        if db: db.close()
        return render_template('keyword_search_results.html', results=None, error=error_msg)
    except Exception as e:
        error_msg = f'Allgemeiner Fehler: {str(e)}'
        print(f"Fehler in /keyword_search_results: {error_msg}")
        if db: db.close()
        return render_template('keyword_search_results.html', results=None, error=error_msg)

# --- Hilfsfunktion zum Extrahieren der PLZ und Finden der Vertriebsnummer ---
def get_plz_and_vertriebsnummer(plz, cursor):
    """Sucht die Vertriebsnummer für eine gegebene PLZ."""
    # Verwende app.logger statt print
    logger = app.logger
    vertriebsnummer = 'N/A' # Standardwert

    if not plz:
        logger.debug(f"get_plz_and_vertriebsnummer: Keine PLZ übergeben.")
        return plz, vertriebsnummer

    try:
        # Finde die postal_code_id für die PLZ
        sql_pc_id = "SELECT postal_code_id FROM postal_code WHERE postal_code = ?" # MODIFIZIERT: code -> postal_code
        logger.debug(f"Führe DB-Abfrage aus: {sql_pc_id} mit PLZ: {plz}")
        cursor.execute(sql_pc_id, (plz,))
        pc_result = cursor.fetchone()
        if pc_result:
            postal_code_id = pc_result['postal_code_id']
            logger.debug(f"Gefundene postal_code_id: {postal_code_id} für PLZ: {plz}")
            # Finde die zugehörige Vertriebsnummer
            sql_vn = """
                SELECT event_gastro_text FROM plz_event_gastro
                WHERE postal_code_id = ? AND event_gastro_text IS NOT NULL AND event_gastro_text != ''
                LIMIT 1
            """
            logger.debug(f"Führe DB-Abfrage aus: {sql_vn[:100]}... mit postal_code_id: {postal_code_id}") # Gekürzte Query loggen
            cursor.execute(sql_vn, (postal_code_id,))
            vn_result = cursor.fetchone()
            if vn_result:
                vertriebsnummer = vn_result['event_gastro_text']
                logger.debug(f"Gefundene Vertriebsnummer: {vertriebsnummer}")
            else:
                logger.debug(f"Keine Vertriebsnummer für postal_code_id {postal_code_id} gefunden.")
        else:
            logger.warning(f"Keine postal_code_id für PLZ {plz} gefunden.")

    except sqlite3.Error as e:
        logger.error(f"DB Fehler beim Suchen der Vertriebsnummer für PLZ {plz}: {e}", exc_info=True)
        # Fehler nicht weiter werfen, N/A wird zurückgegeben
    except Exception as e:
         logger.error(f"Allgemeiner Fehler beim Suchen der Vertriebsnummer für PLZ {plz}: {e}", exc_info=True)

    return plz, vertriebsnummer

# --- Hilfsfunktion zum Trennen der Adresse ---
def split_address(formatted_address):
    """Versucht, Straße+Nr und PLZ+Stadt zu trennen."""
    street_number = ''
    plz_city = ''

    if not formatted_address:
        return street_number, plz_city

    # Annahme: Adresse ist oft 'Straße Nr, PLZ Stadt' oder 'Straße Nr\\nPLZ Stadt'
    parts = formatted_address.replace('\\n', ',').split(',')
    if len(parts) >= 2:
        street_number = parts[0].strip()
        plz_city = parts[1].strip()
        # Wenn die PLZ im ersten Teil ist (selten, aber möglich)
        if re.search(r'\d{5}', street_number) and not re.search(r'\d{5}', plz_city):
            street_number, plz_city = plz_city, street_number
    elif len(parts) == 1:
        # Versuche, anhand der PLZ zu trennen
        match = re.search(r'(.*?)(\d{5}\s+.*)', formatted_address)
        if match:
            street_number = match.group(1).strip().rstrip(',') # Komma am Ende entfernen
            plz_city = match.group(2).strip()
        else:
             # Fallback: Alles in PLZ/Stadt, da Trennung unsicher
             plz_city = formatted_address.strip()

    return street_number, plz_city

@app.route('/export_keyword_results', methods=['POST'])
def export_keyword_results():
    """Exportiert ausgewählte Keyword-Suchergebnisse als Excel."""
    conn = None # Initialize conn outside try
    logger = app.logger # Logger holen
    try:
        request_data = request.get_json()
        selected_place_ids = request_data.get('place_ids', [])

        if not selected_place_ids:
            logger.warning("Export abgebrochen: Keine Orte ausgewählt.")
            return jsonify({'error': 'Keine Orte für den Export ausgewählt.'}), 400

        # Konvertiere IDs zu Strings für die Abfrage
        place_ids_tuple = tuple(map(str, selected_place_ids))
        logger.info(f"Exportiere Daten für Place IDs: {place_ids_tuple}")

        # --- Datenbankabfrage --- #
        export_data = [] # Initialisiere hier
        try:
            conn = get_db()
            cursor = conn.cursor()
            logger.info("Datenbankverbindung für Export geöffnet.")

            # Daten für die ausgewählten Orte abrufen
            query = f"""
                SELECT
                    p.place_id,
                    p.name AS place_name,
                    p.display_name AS place_display_name,
                    p.formatted_address,
                    p.phone_number,
                    p.website_uri,
                    p.google_maps_uri,
                    c.name AS city_name,
                    p.postal_code
                FROM place p
                JOIN city c ON p.city_id = c.city_id
                WHERE p.place_id IN ({','.join(['?'] * len(place_ids_tuple))})
            """
            logger.debug(f"Führe Hauptabfrage aus: {query} mit Parametern: {place_ids_tuple}")
            cursor.execute(query, place_ids_tuple)
            places_data = cursor.fetchall()
            logger.info(f"{len(places_data)} Orte aus DB geholt.")

            if not places_data:
                 logger.warning("Keine Daten für die ausgewählten Orte in der DB gefunden.")
                 # Hier nicht direkt returnen, conn muss noch geschlossen werden
            else:
                # Hole zusätzliche Daten für jeden Ort
                for place_row in places_data:
                    place_id = place_row['place_id']
                    place_dict = dict(place_row)
                    logger.debug(f"Verarbeite zusätzlichen Daten für Place ID: {place_id}")

                    # 1. Neueste Bewertung holen
                    logger.debug(f"Hole Bewertung für {place_id}")
                    cursor.execute("""
                        SELECT rating, user_rating_count FROM rating_history
                        WHERE place_id = ?
                        ORDER BY timestamp DESC LIMIT 1
                    """, (place_id,))
                    rating_result = cursor.fetchone()
                    place_dict['rating'] = rating_result['rating'] if rating_result else None
                    place_dict['user_rating_count'] = rating_result['user_rating_count'] if rating_result else None

                    # 2. Zugehörige Suchbegriffe holen
                    logger.debug(f"Hole Keywords für {place_id}")
                    cursor.execute("""
                        SELECT DISTINCT st.name FROM search_term st
                        JOIN place_search ps ON st.term_id = ps.term_id
                        WHERE ps.place_id = ?
                    """, (place_id,))
                    keyword_results = cursor.fetchall()
                    place_dict['keywords'] = ', '.join([kw['name'] for kw in keyword_results])

                    # 3. Adresse bleibt zusammen (formatted_address wird direkt verwendet)
                    # street_number, plz_city = split_address(place_dict['formatted_address'])
                    # place_dict['street_number'] = street_number
                    # place_dict['plz_city'] = plz_city

                    # 4. PLZ extrahieren (aus DB-Spalte) und Vertriebsnummer suchen
                    db_plz = place_dict.get('postal_code') # Hole PLZ aus der DB-Spalte
                    place_dict['extracted_plz'] = db_plz if db_plz else '' # Speichere PLZ für Export
                    logger.debug(f"Suche Vertriebsnummer für PLZ: {db_plz} (Place ID: {place_id})")
                    # Übergebe nur die extrahierte PLZ an die Mapping-Funktion
                    _, vertriebsnummer = get_plz_and_vertriebsnummer(db_plz, cursor)
                    place_dict['vertriebsnummer'] = vertriebsnummer

                    export_data.append(place_dict)

                logger.info("Zusätzliche Daten für alle Orte erfolgreich geholt.")

        except sqlite3.Error as db_err:
            # Log the specific DB error
            error_msg = f"DB Fehler in /export_keyword_results (Datenabfrage): {db_err}"
            logger.error(error_msg, exc_info=True) # Log with traceback
            # Fehler weitergeben, damit er im äußeren try/except gefangen wird
            raise db_err
        finally:
            # Ensure connection is closed if opened
            if conn:
                conn.close()
                logger.info("Datenbankverbindung nach Export-Datenabfrage geschlossen.")

        # --- DataFrame und Excel-Verarbeitung --- #
        # Dieser Block wird nur ausgeführt, wenn kein DB-Fehler aufgetreten ist
        try:
            if not export_data:
                logger.warning("Export abgebrochen: Keine Daten zum Verarbeiten nach DB-Abfrage.")
                return jsonify({'error': 'Keine Daten zum Exportieren vorhanden.'}), 404

            logger.info("Erstelle DataFrame für Export...")
            df_export = pd.DataFrame(export_data)

            # Spalten auswählen und umbenennen
            export_columns_map = {
                'keywords': 'Suchbegriff',
                'city_name': 'Stadt',
                'place_display_name': 'Name',
                'rating': 'Gesamtbewertung',
                'user_rating_count': 'Anzahl Bewertungen',
                'formatted_address': 'Adresse', # Geändert: Komplette Adresse
                'extracted_plz': 'PLZ',        # Neu: Nur PLZ
                'phone_number': 'Telefonnummer',
                'website_uri': 'Website',
                'google_maps_uri': 'Google Maps Link',
                'vertriebsnummer': 'Vertriebsnummer'
            }

            # Robustere Spaltenbehandlung
            if 'place_display_name' in df_export.columns and 'place_name' in df_export.columns:
                 df_export['place_display_name'] = df_export['place_display_name'].fillna(df_export['place_name'])
            elif 'place_name' in df_export.columns:
                 df_export['place_display_name'] = df_export['place_name']
            else:
                 df_export['place_display_name'] = 'Name nicht verfügbar'

            final_export_columns = []
            for source_col, target_col in export_columns_map.items():
                if source_col not in df_export.columns:
                    logger.warning(f"Fehlende Quellspalte '{source_col}' für Export. Füge leere Spalte '{target_col}' hinzu.")
                    # Füge die fehlende Spalte hinzu (könnte z.B. 'postal_code' sein, wenn es in DB fehlt)
                    df_export[source_col] = None if source_col != 'extracted_plz' else ''
                final_export_columns.append(source_col)

            df_export_final = df_export[final_export_columns].rename(columns=export_columns_map)
            logger.info("DataFrame für Export vorbereitet.")

            # Datentypen anpassen
            logger.info("Passe Datentypen für Excel an...")
            df_export_final['Gesamtbewertung'] = pd.to_numeric(df_export_final['Gesamtbewertung'], errors='coerce')
            df_export_final['Anzahl Bewertungen'] = pd.to_numeric(df_export_final['Anzahl Bewertungen'], errors='coerce').fillna(0).astype(int)
            df_export_final['PLZ'] = df_export_final['PLZ'].astype(str) # Sicherstellen, dass PLZ als Text behandelt wird

            # Excel-Datei im Speicher erstellen
            logger.info("Erstelle Excel-Datei im Speicher...")
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_export_final.to_excel(writer, index=False, sheet_name='Keyword Ergebnisse')
                workbook = writer.book
                worksheet = writer.sheets['Keyword Ergebnisse']
                logger.info("Daten in Excel geschrieben, starte Formatierung...")

                # Formatierungen anwenden
                header_font = Font(bold=True)
                centered_alignment = Alignment(horizontal='center', vertical='center')
                right_alignment = Alignment(horizontal='right', vertical='center')
                left_alignment = Alignment(horizontal='left', vertical='center')
                number_format_decimal = '0.0'
                number_format_integer = '#,##0'

                for col_idx, column_name in enumerate(df_export_final.columns, 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.alignment = centered_alignment
                    column_letter = openpyxl.utils.get_column_letter(col_idx)

                    logger.debug(f"Formatiere Spalte: {column_name} (Index: {col_idx})")

                    # Standardbreite
                    worksheet.column_dimensions[column_letter].width = 18

                    # Spezifische Breiten und Formate
                    if column_name == 'Name':
                         worksheet.column_dimensions[column_letter].width = 30
                         for row_idx in range(2, worksheet.max_row + 1):
                             worksheet[f"{column_letter}{row_idx}"].alignment = left_alignment
                    elif column_name == 'Adresse': # Geändert
                         worksheet.column_dimensions[column_letter].width = 45 # Breiter für komplette Adresse
                         for row_idx in range(2, worksheet.max_row + 1):
                             worksheet[f"{column_letter}{row_idx}"].alignment = left_alignment
                    elif column_name == 'PLZ': # Neu
                         worksheet.column_dimensions[column_letter].width = 10
                         for row_idx in range(2, worksheet.max_row + 1):
                             worksheet[f"{column_letter}{row_idx}"].alignment = left_alignment # Oder centered_alignment?
                             # Sicherstellen, dass PLZ als Text formatiert wird
                             worksheet[f"{column_letter}{row_idx}"].number_format = '@'
                    elif column_name in ['Website', 'Google Maps Link']:
                         worksheet.column_dimensions[column_letter].width = 40
                         for row_idx in range(2, worksheet.max_row + 1):
                             worksheet[f"{column_letter}{row_idx}"].alignment = left_alignment
                    elif column_name == 'Suchbegriff':
                         worksheet.column_dimensions[column_letter].width = 25
                         for row_idx in range(2, worksheet.max_row + 1):
                              worksheet[f"{column_letter}{row_idx}"].alignment = left_alignment
                    elif column_name == 'Stadt':
                         worksheet.column_dimensions[column_letter].width = 20
                         for row_idx in range(2, worksheet.max_row + 1):
                              worksheet[f"{column_letter}{row_idx}"].alignment = left_alignment
                    elif column_name == 'Gesamtbewertung':
                         worksheet.column_dimensions[column_letter].width = 15
                         for row_idx in range(2, worksheet.max_row + 1):
                             target_cell = worksheet[f"{column_letter}{row_idx}"]
                             if isinstance(target_cell.value, (int, float)):
                                 target_cell.number_format = number_format_decimal
                             target_cell.alignment = right_alignment
                    elif column_name == 'Anzahl Bewertungen':
                         worksheet.column_dimensions[column_letter].width = 18
                         for row_idx in range(2, worksheet.max_row + 1):
                             target_cell = worksheet[f"{column_letter}{row_idx}"]
                             if isinstance(target_cell.value, (int, float)):
                                 target_cell.number_format = number_format_integer
                             target_cell.alignment = right_alignment
                    elif column_name in ['Telefonnummer', 'Vertriebsnummer']:
                         worksheet.column_dimensions[column_letter].width = 20
                         for row_idx in range(2, worksheet.max_row + 1):
                              worksheet[f"{column_letter}{row_idx}"].alignment = left_alignment

            output.seek(0)
            logger.info("Excel-Datei erfolgreich erstellt und formatiert.")

            today = datetime.now().strftime("%Y%m%d")
            filename = f"Keyword_Ergebnisse_{today}.xlsx"

            # Datei zum Download senden
            logger.info(f"Sende Datei '{filename}' zum Download.")
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )

        except Exception as proc_err:
             # Log any error during DataFrame or Excel processing
             error_msg = f"Fehler bei DataFrame/Excel Verarbeitung in /export_keyword_results: {proc_err}"
             logger.error(error_msg, exc_info=True) # Log with traceback
             return jsonify({'error': f'Fehler bei der Dateierstellung: {proc_err}'}), 500

    except sqlite3.Error as db_err_outer: # Fange DB-Fehler hier außen auch
        # Dieser Block fängt DB-Fehler, die im inneren try nicht gefangen wurden (sollte nicht passieren)
        # oder wenn der Fehler *vor* dem inneren DB-Try auftritt (unwahrscheinlich hier)
        error_msg = f"Äußerer DB Fehler in /export_keyword_results: {db_err_outer}"
        logger.error(error_msg, exc_info=True)
        # Stelle sicher, dass conn geschlossen ist, falls es doch geöffnet wurde
        if conn: conn.close()
        return jsonify({'error': f'Datenbankfehler: {db_err_outer}'}), 500
    except Exception as e:
         # Catch-all for unexpected errors (e.g., JSON parsing)
         error_msg = f"Unerwarteter Fehler in /export_keyword_results: {e}"
         logger.error(error_msg, exc_info=True) # Log with traceback
         # Ensure conn is closed if it exists and the error happened before the specific finally block
         if conn: conn.close()
         return jsonify({'error': f'Interner Serverfehler: {e}'}), 500

@app.route('/api_metrics', methods=['GET'])
def api_metrics():
    """Zeigt die Google Maps API Nutzungsstatistiken für den aktuellen Monat an."""
    try:
        # Prüfe, ob Service Account Datei existiert
        service_account_file = os.environ.get('GOOGLE_CLOUD_SERVICE_ACCOUNT_FILE', 'service-account-key.json')
        
        if not os.path.exists(service_account_file):
            return jsonify({
                'error': 'Service Account nicht konfiguriert',
                'total_requests': 0,
                'warning_message': [
                    "Achtung: Es können nur 1000 Abfragen im Monat durchgeführt werden!",
                    "Ein Durchgang sind 400 Abfragen.",
                    "API Metriken nicht verfügbar (Service Account fehlt)."
                ]
            }), 200  # Return 200 with default data instead of error
        
        # Projekt-ID aus Umgebungsvariablen
        project_id = os.environ.get('GOOGLE_CLOUD_PROJECT_ID')
        if not project_id:
            return jsonify({
                'error': 'Projekt-ID nicht konfiguriert',
                'total_requests': 0,
                'warning_message': [
                    "Achtung: Es können nur 1000 Abfragen im Monat durchgeführt werden!",
                    "Ein Durchgang sind 400 Abfragen.",
                    "API Metriken nicht verfügbar (Projekt-ID fehlt)."
                ]
            }), 200
        
        # Authentifizierung
        credentials = service_account.Credentials.from_service_account_file(service_account_file)
        
        # Client erstellen
        client = monitoring_v3.MetricServiceClient(credentials=credentials)
        project_name = f"projects/{project_id}"
        
        # Zeitraum für aktuellen Kalendermonat definieren
        now = datetime.now()
        # Erster Tag des aktuellen Monats
        start_of_month = datetime(now.year, now.month, 1)
        
        # Zeitintervall erstellen
        interval = monitoring_v3.TimeInterval(
            {
                "start_time": start_of_month,
                "end_time": now,
            }
        )
        
        # Metriken für Places API (New) abfragen
        try:
            results = client.list_time_series(
                request={
                    "name": project_name,
                    "filter": 'metric.type="serviceruntime.googleapis.com/api/request_count" AND resource.label.service="places.googleapis.com"',
                    "interval": interval,
                    "view": monitoring_v3.ListTimeSeriesRequest.TimeSeriesView.FULL,
                }
            )
            
            # Ergebnisse verarbeiten
            total_requests = 0
            
            for time_series in results:
                for point in time_series.points:
                    total_requests += point.value.int64_value
            
            # Formatiere die Antwort
            response_data = {
                'total_requests': total_requests,
                'warning_message': [
                    "Achtung: Es können nur 1000 Abfragen im Monat durchgeführt werden!",
                    "Ein Durchgang sind 400 Abfragen.",
                    f"Diesen Monat wurden bereits {total_requests} Abfragen durchgeführt."
                ]
            }
            
            return jsonify(response_data)
            
        except Exception as monitoring_error:
            logger = app.logger
            logger.error(f"Fehler beim Abrufen der Cloud Monitoring Metriken: {monitoring_error}", exc_info=True)
            return jsonify({
                'error': 'Metriken konnten nicht abgerufen werden',
                'total_requests': 0,
                'warning_message': [
                    "Achtung: Es können nur 1000 Abfragen im Monat durchgeführt werden!",
                    "Ein Durchgang sind 400 Abfragen.",
                    "API Metriken temporär nicht verfügbar."
                ]
            }), 200
            
    except Exception as e:
        logger = app.logger
        logger.error(f"Allgemeiner Fehler in /api_metrics: {e}", exc_info=True)
        return jsonify({
            'error': 'Konfigurationsfehler',
            'total_requests': 0,
            'warning_message': [
                "Achtung: Es können nur 1000 Abfragen im Monat durchgeführt werden!",
                "Ein Durchgang sind 400 Abfragen.",
                "API Metriken nicht verfügbar."
            ]
        }), 200

@app.route('/api_metrics_page')
def api_metrics_page():
    """Rendert die HTML-Seite für die API Metriken."""
    return render_template('api_metrics.html')

if __name__ == '__main__':
    # Stelle sicher, dass der API Key geladen wird
    if not API_KEY or API_KEY == 'DEIN_API_KEY':
        print("WARNUNG: GOOGLE_MAPS_API_KEY nicht gesetzt oder ist Platzhalter.")
        print("           Die Google Places Suche wird nicht funktionieren.")
        # Hier könnte man die App auch beenden oder den Key abfragen
        # exit(1) # Beenden, wenn Key zwingend erforderlich

    # Führe init_db nur aus, wenn es der Hauptprozess ist (nicht der Reloader-Subprozess)
    # Die folgende Bedingung und der Aufruf werden entfernt
    # if os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
    #     init_db() # Initialisiere die Datenbank nur einmal

    app.run(debug=True, threaded=True) # threaded=True ist wichtig für Hintergrundsuche und SSE
